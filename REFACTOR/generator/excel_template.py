"""
Excel Template Generator
Generates an Excel template file with example data and instructions.
All examples are generic - no application-specific hard-coding.
"""
import pandas as pd
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def generate_excel_template(output_path: Path, include_examples: bool = True) -> Path:
    """
    Generate Excel template file with example data and instructions.
    
    Args:
        output_path: Path where template will be saved
        include_examples: Whether to include example rows (default: True)
        
    Returns:
        Path to generated template file
    """
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create main data sheet
    ws_data = wb.create_sheet("Test Steps", 0)
    
    # Define column headers
    headers = [
        'Step',
        'URL',
        'XPath',
        'Object Type',
        'Action',
        'Functions',
        'Text Value',
        'Wait Time',
        'Optional'
    ]
    
    # Write headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add example rows if requested
    if include_examples:
        example_rows = _get_example_rows()
        for row_num, row_data in enumerate(example_rows, start=2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_data.cell(row=row_num, column=col_num, value=value)
                # Style example rows differently
                if row_num <= len(example_rows) + 1:
                    cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        # Set width based on header length and content
        max_length = len(header)
        if include_examples:
            for row_num in range(2, len(example_rows) + 2):
                cell_value = ws_data.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        ws_data.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Create instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 1)
    _add_instructions(ws_instructions)
    
    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return output_path


def _get_example_rows() -> list:
    """
    Get example rows for template.
    All examples are generic - no application-specific values.
    
    Returns:
        List of example row data
    """
    return [
        # Step 1: Navigate
        [
            1,
            'https://example.com',
            'N/A',
            '',
            'navigate',
            '',
            '',
            '',
            'false'
        ],
        # Step 2: Wait
        [
            2,
            'N/A',
            'N/A',
            '',
            'wait',
            '',
            '',
            2000,
            'false'
        ],
        # Step 3: Click button
        [
            3,
            'https://example.com',
            "//button[@type='submit']",
            'button',
            'click',
            '',
            '',
            '',
            'false'
        ],
        # Step 4: Fill input
        [
            4,
            'https://example.com',
            "//input[@type='text']",
            'input',
            'fill',
            '',
            'Sample Text',
            '',
            'false'
        ],
        # Step 5: Fill password
        [
            5,
            'https://example.com',
            "//input[@type='password']",
            'input',
            'fill',
            '',
            'password123',
            '',
            'false'
        ],
        # Step 6: Fill with TOTP function
        [
            6,
            'https://example.com',
            "//input[@class='otp-input']",
            'input',
            'fill',
            'TOTP',
            '',
            '',
            'false'
        ],
        # Step 7: Click dropdown option
        [
            7,
            'https://example.com',
            "//select[@id='dropdown']//option[text()='Option 1']",
            'dropdown',
            'click',
            '',
            '',
            '',
            'false'
        ],
        # Step 8: Optional step
        [
            8,
            'https://example.com',
            "//button[contains(text(), 'Optional')]",
            'button',
            'click',
            '',
            '',
            '',
            'true'
        ],
        # Step 9: Verify element
        [
            9,
            'https://example.com',
            "//div[@class='success-message']",
            'div',
            'verify',
            '',
            '',
            '',
            'false'
        ],
    ]


def _add_instructions(ws):
    """
    Add instructions to the instructions sheet.
    
    Args:
        ws: Worksheet to add instructions to
    """
    instructions = [
        ["Excel Test Case Template - Instructions", ""],
        ["", ""],
        ["Column Descriptions:", ""],
        ["", ""],
        ["Step", "Step number or identifier (e.g., 1, 2, 3 or '1a', '2b')"],
        ["URL", "Page URL to navigate to. Use 'N/A' if no navigation needed."],
        ["XPath", "XPath selector for the element. Use 'N/A' if not applicable."],
        ["Object Type", "Type of element: button, input, link, dropdown, checkbox, radio, text, div, span, a"],
        ["Action", "Action to perform: navigate, click, fill, verify, wait"],
        ["Functions", "Special functions: TOTP (for one-time password generation)"],
        ["Text Value", "Text to fill (for 'fill' actions). Use '${TIMESTAMP}' for dynamic timestamp."],
        ["Wait Time", "Wait time in milliseconds (for 'wait' actions)"],
        ["Optional", "Whether step is optional: true/false"],
        ["", ""],
        ["Action Types:", ""],
        ["", ""],
        ["navigate", "Navigate to the URL specified in the URL column"],
        ["click", "Click the element specified by XPath"],
        ["fill", "Fill the input element with Text Value"],
        ["verify", "Verify that the element exists"],
        ["wait", "Wait for the specified Wait Time (in milliseconds)"],
        ["", ""],
        ["Special Functions:", ""],
        ["", ""],
        ["TOTP", "Automatically generate and fill a time-based one-time password"],
        ["", ""],
        ["Notes:", ""],
        ["", ""],
        ["- All XPaths should be valid XPath expressions"],
        ["- URLs must start with http:// or https://"],
        ["- Use 'N/A' for optional fields that don't apply"],
        ["- Optional steps will be skipped if they fail"],
        ["- Text Value supports ${TIMESTAMP} placeholder for dynamic timestamps"],
        ["", ""],
        ["Example XPaths:", ""],
        ["", ""],
        ["//button[@type='submit']", "Button with type='submit'"],
        ["//input[@type='text']", "Text input field"],
        ["//a[contains(text(), 'Login')]", "Link containing 'Login' text"],
        ["(//div[@class='modal'])[1]", "First div with class='modal'"],
        ["//select[@id='dropdown']//option[text()='Option']", "Dropdown option"],
    ]
    
    # Style title
    title_cell = ws.cell(row=1, column=1, value=instructions[0][0])
    title_cell.font = Font(bold=True, size=14)
    
    # Write instructions
    for row_num, row_data in enumerate(instructions, start=1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif row_num <= 3:
                cell.font = Font(bold=True)
            elif value and value.endswith(':'):
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60


def get_template_path(template_name: str = "test_case_template.xlsx", project_root: Path = None) -> Path:
    """
    Get the default path for template file.
    
    Args:
        template_name: Name of template file
        project_root: Optional project root path (defaults to current directory)
        
    Returns:
        Path to template file
    """
    # Use project root if provided, otherwise use current directory
    if project_root is None:
        project_root = Path.cwd()
    
    # Default to storage/excel_files/templates/ directory
    template_dir = project_root / "storage" / "excel_files" / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)
    return template_dir / template_name

