"""API Routes"""
from flask import Blueprint, request, jsonify, current_app, send_file, render_template
import json
import sys
import asyncio
import time
from pathlib import Path
from datetime import datetime
import threading

sys.path.insert(0, str(Path(__file__).parent.parent))
from agent.core.agent import Agent
from utils.element_registry import get_registry

# Excel generator imports
from REFACTOR.generator.excel_generator_ts import generate_playwright_ts_from_excel
from REFACTOR.generator.excel_validator import validate_excel_file, get_validation_summary
from REFACTOR.generator.excel_template import generate_excel_template, get_template_path
from REFACTOR.generator.excel_registry_helper import extract_elements_from_excel, compare_with_registry
import uuid
import pandas as pd
import shutil

bp = Blueprint('api', __name__)
active_executions = {}
active_excel_generations = {}


def extract_file_upload_paths_from_excel(excel_path: Path) -> list:
    """
    Extract file upload paths from Excel file by parsing Functions column.
    Supports both file paths and folder paths (ending with /).
    
    Args:
        excel_path: Path to Excel file
        
    Returns:
        List of file/folder paths (e.g., ['storage/test_files/file1.tsv', 'storage/test_files/cds/'])
    """
    file_paths = []
    try:
        df = pd.read_excel(excel_path)
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        if 'functions' not in df.columns:
            return file_paths
        
        for idx, row in df.iterrows():
            functions = str(row.get('functions', '')).strip() if pd.notna(row.get('functions')) else ''
            
            if functions and 'file upload' in functions.lower():
                # Parse file path from functions: "File Upload:storage/test_files/file.tsv" or "File Upload:storage/test_files/cds/"
                if ':' in functions:
                    parts = functions.split(':')
                    if len(parts) >= 2:
                        # Get path after "File Upload:"
                        path_part = ':'.join(parts[1:]).strip()
                        if path_part:
                            # Normalize folder paths to end with /
                            if path_part.endswith('/'):
                                path_part = path_part.rstrip('/') + '/'
                            if path_part not in file_paths:
                                file_paths.append(path_part)
    
    except Exception as e:
        print(f"⚠️  Error extracting file upload paths from Excel: {e}")
    
    return file_paths


def upload_test_files_to_server(test_files: list, project_root: Path, referenced_paths: list = None) -> dict:
    """
    Upload test files to server, creating folder structure as needed.
    Automatically renames uploaded files to match Excel-referenced paths when there's a 1:1 match.
    
    Args:
        test_files: List of file objects from request.files.getlist('test_files')
        project_root: Project root directory
        referenced_paths: List of file paths referenced in Excel (e.g., ['storage/test_files/file.tsv'])
        
    Returns:
        Dict with upload results: {'uploaded': [...], 'errors': [...], 'renamed': [...]}
    """
    results = {'uploaded': [], 'errors': [], 'renamed': []}
    test_files_dir = project_root / 'storage' / 'test_files'
    
    # Create test_files directory if it doesn't exist
    test_files_dir.mkdir(parents=True, exist_ok=True)
    
    # Filter out empty filenames
    valid_files = [f for f in test_files if f.filename]
    
    # Check if referenced path is a folder (ends with /)
    if referenced_paths and len(referenced_paths) == 1:
        referenced_path = referenced_paths[0]
        is_folder_path = referenced_path.endswith('/')
        
        if is_folder_path:
            # Folder path: save all uploaded files to this folder
            # e.g., "storage/test_files/cds/" -> save all files to storage/test_files/cds/
            if referenced_path.startswith('storage/test_files/'):
                folder_relative = referenced_path.replace('storage/test_files/', '')
            elif referenced_path.startswith('test_files/'):
                folder_relative = referenced_path.replace('test_files/', '')
            else:
                # Remove leading storage/ if present
                folder_relative = referenced_path.replace('storage/', '', 1) if referenced_path.startswith('storage/') else referenced_path
            
            # Remove trailing /
            folder_relative = folder_relative.rstrip('/')
            
            # Create target folder
            target_folder = test_files_dir / folder_relative
            target_folder.mkdir(parents=True, exist_ok=True)
            
            # Save all uploaded files to this folder
            for file in valid_files:
                try:
                    target_path = target_folder / file.filename
                    file.save(str(target_path))
                    relative_path = str(target_path.relative_to(project_root))
                    results['uploaded'].append({
                        'filename': file.filename,
                        'server_path': relative_path,
                        'full_path': str(target_path),
                        'saved_to_folder': relative_path
                    })
                    print(f"✅ Uploaded test file to folder: {file.filename} -> {relative_path}")
                except Exception as e:
                    error_msg = f"Failed to upload {file.filename}: {str(e)}"
                    results['errors'].append(error_msg)
                    print(f"❌ {error_msg}")
            
            return results
        
        elif len(valid_files) == 1:
            # Single file path: auto-rename to match Excel reference
            file = valid_files[0]
            
            # Extract filename from referenced path (e.g., "storage/test_files/file.tsv" -> "file.tsv")
            # But keep the directory structure
            if referenced_path.startswith('storage/test_files/'):
                target_relative = referenced_path.replace('storage/test_files/', '')
            elif referenced_path.startswith('test_files/'):
                target_relative = referenced_path.replace('test_files/', '')
            else:
                # Use referenced path as-is, but ensure it's relative to test_files_dir
                target_relative = referenced_path.split('/')[-1] if '/' in referenced_path else referenced_path
            
            # Create directory structure if needed
            if '/' in target_relative:
                parts = target_relative.split('/')
                subdir = test_files_dir / '/'.join(parts[:-1])
                subdir.mkdir(parents=True, exist_ok=True)
                target_path = subdir / parts[-1]
            else:
                target_path = test_files_dir / target_relative
            
            try:
                file.save(str(target_path))
                relative_path = str(target_path.relative_to(project_root))
                results['uploaded'].append({
                    'filename': file.filename,
                    'server_path': relative_path,
                    'full_path': str(target_path),
                    'renamed_to': relative_path
                })
                results['renamed'].append({
                    'original': file.filename,
                    'renamed_to': relative_path
                })
                print(f"✅ Uploaded and renamed test file: {file.filename} -> {relative_path} (matched Excel reference)")
            except Exception as e:
                error_msg = f"Failed to upload {file.filename}: {str(e)}"
                results['errors'].append(error_msg)
                print(f"❌ {error_msg}")
            
            return results
    
    # Multiple files or no referenced paths - upload with original names
    for file in valid_files:
        try:
            # Extract relative path from filename if it contains directory structure
            # e.g., "test_files/subfolder/file.tsv" -> create subfolder structure
            filename = file.filename
            if '/' in filename or '\\' in filename:
                # Handle both forward and backslash separators
                parts = filename.replace('\\', '/').split('/')
                # Remove 'test_files' prefix if present (we're already in test_files_dir)
                if parts[0].lower() == 'test_files':
                    parts = parts[1:]
                
                # Create subdirectory structure
                if len(parts) > 1:
                    subdir = test_files_dir / '/'.join(parts[:-1])
                    subdir.mkdir(parents=True, exist_ok=True)
                    target_path = subdir / parts[-1]
                else:
                    target_path = test_files_dir / parts[0]
            else:
                # Simple filename, save directly in test_files_dir
                target_path = test_files_dir / filename
            
            # Save file
            file.save(str(target_path))
            relative_path = str(target_path.relative_to(project_root))
            results['uploaded'].append({
                'filename': filename,
                'server_path': relative_path,
                'full_path': str(target_path)
            })
            print(f"✅ Uploaded test file: {filename} -> {relative_path}")
        
        except Exception as e:
            error_msg = f"Failed to upload {file.filename}: {str(e)}"
            results['errors'].append(error_msg)
            print(f"❌ {error_msg}")
    
    return results


@bp.route('/execute', methods=['POST'])
def execute_story():
    try:
        data = request.get_json()
        story = data.get('story', '').strip()
        
        if not story:
            return jsonify({'error': 'Story required'}), 400
        
        agent = Agent()
        execution_id = agent.context.execution_id
        
        # Get project root before threading
        project_root = current_app.config['PROJECT_ROOT']
        
        active_executions[execution_id] = {
            'agent': agent,
            'status': 'running',
            'story': story,
            'started_at': datetime.now().isoformat()
        }
        
        def run_execution():
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                results = loop.run_until_complete(agent.execute_story(story))
                
                # Use project_root from closure
                results_dir = project_root / 'storage' / 'executions'
                results_dir.mkdir(parents=True, exist_ok=True)
                
                results_file = results_dir / f'{execution_id}.json'
                with open(results_file, 'w') as f:
                    json.dump(results, f, indent=2)
                
                active_executions[execution_id]['status'] = results['status']
                active_executions[execution_id]['results'] = results
            except Exception as e:
                import traceback
                import logging
                logger = logging.getLogger(__name__)
                active_executions[execution_id]['status'] = 'error'
                active_executions[execution_id]['error'] = str(e)
                logger.error(f"❌ Error in run_execution: {e}")
                logger.error(traceback.format_exc())
                print(f"Error in run_execution: {e}")
                print(traceback.format_exc())
        
        thread = threading.Thread(target=run_execution, daemon=True)
        thread.start()
        
        return jsonify({
            'execution_id': execution_id,
            'status': 'started'
        }), 202
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<execution_id>/status', methods=['GET'])
def get_execution_status(execution_id):
    if execution_id in active_executions:
        exec_data = active_executions[execution_id]
        response = {
            'execution_id': execution_id,
            'status': exec_data['status'],
            'story': exec_data['story'],
            'started_at': exec_data['started_at']
        }
        if 'error' in exec_data:
            response['error'] = exec_data['error']
        if 'results' in exec_data:
            results = exec_data['results']
            response.update({
                'actions_count': len(results.get('actions_taken', [])),
                'screenshots_count': len(results.get('screenshots', [])),
                'summary': results.get('summary'),
                'error': results.get('error')
            })
        elif 'agent' in exec_data:
            # Get live progress from agent context while running
            agent = exec_data['agent']
            if agent and hasattr(agent, 'context'):
                response['actions_count'] = len(agent.context.actions_taken)
                response['screenshots_count'] = len(agent.context.screenshots)
        return jsonify(response), 200
    
    project_root = current_app.config['PROJECT_ROOT']
    results_file = project_root / 'storage' / 'executions' / f'{execution_id}.json'
    
    if results_file.exists():
        with open(results_file) as f:
            results = json.load(f)
        return jsonify({
            'execution_id': execution_id,
            'status': results['status'],
            'story': results['story'],
            'actions_count': len(results.get('actions_taken', [])),
            'screenshots_count': len(results.get('screenshots', [])),
            'summary': results.get('summary'),
            'error': results.get('error')
        }), 200
    
    return jsonify({'error': 'Not found'}), 404


@bp.route('/executions/<execution_id>/results', methods=['GET'])
def get_execution_results(execution_id):
    # ALWAYS load from file first (to get latest Playwright data if it was added)
    project_root = current_app.config['PROJECT_ROOT']
    results_file = project_root / 'storage' / 'executions' / f'{execution_id}.json'
    
    if results_file.exists():
        with open(results_file) as f:
            return jsonify(json.load(f)), 200
    
    # Fallback to live results from active executions (for in-progress tests)
    if execution_id in active_executions:
        exec_data = active_executions[execution_id]
        if 'results' in exec_data:
            return jsonify(exec_data['results']), 200
        elif 'agent' in exec_data:
            # Return partial results while running
            agent = exec_data['agent']
            return jsonify({
                'execution_id': execution_id,
                'status': exec_data['status'],
                'story': exec_data['story'],
                'actions_taken': [],
                'screenshots': []
            }), 200
    
    return jsonify({'error': 'Not found'}), 404


@bp.route('/executions/<execution_id>/validation-results', methods=['GET'])
def get_validation_results(execution_id):
    """Get validation results JSON file for an execution"""
    project_root = current_app.config['PROJECT_ROOT']
    validation_file = project_root / 'storage' / 'validation_results' / f'{execution_id}.json'
    
    if validation_file.exists():
        with open(validation_file) as f:
            return jsonify(json.load(f)), 200
    
    return jsonify({'error': 'Validation results not found', 'validations': []}), 404


@bp.route('/executions/<execution_id>/validation-results-excel', methods=['GET'])
def get_validation_results_excel(execution_id):
    """Get validation results as Excel file for an execution"""
    try:
        from flask import send_file
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        project_root = current_app.config['PROJECT_ROOT']
        validation_file = project_root / 'storage' / 'validation_results' / f'{execution_id}.json'
        
        if not validation_file.exists():
            return jsonify({'error': 'Validation results not found'}), 404
        
        # Load JSON data
        with open(validation_file) as f:
            validation_data = json.load(f)
        
        # Create Excel workbook in memory
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        validations = validation_data.get('validations', [])
        if not validations:
            return jsonify({'error': 'No validation data found'}), 404
        
        # ============================================================================
        # Sheet 1: Summary
        # ============================================================================
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Validation Summary"])
        ws_summary.append([])
        ws_summary.append(["Total Validations", len(validations)])
        
        total_mismatches = sum(len(v.get('mismatches', [])) for v in validations)
        total_matches = sum(len(v.get('matches', [])) for v in validations)
        ws_summary.append(["Total Mismatches", total_mismatches])
        ws_summary.append(["Total Matches", total_matches])
        ws_summary.append([])
        
        # Per-validation summary
        ws_summary.append(["Step", "Tab Name", "Type", "Status", "Mismatches", "Matches", "Timestamp"])
        for v in validations:
            step = v.get('step', '')
            tab_name = v.get('webTabName', '') or v.get('excelTabName', '')
            val_type = v.get('validationType', 'table')
            mismatches_count = len(v.get('mismatches', []))
            matches_count = len(v.get('matches', []))
            status = "✅ Passed" if mismatches_count == 0 else "❌ Failed"
            timestamp = v.get('timestamp', '')
            
            # For data_view, check summary
            if val_type == 'data_view':
                summary = v.get('summary', {})
                passed = summary.get('passed', 0)
                failed = summary.get('failed', 0)
                status = f"✅ {passed}/{summary.get('totalNodeTypes', 0)} Passed" if failed == 0 else f"❌ {failed}/{summary.get('totalNodeTypes', 0)} Failed"
            
            ws_summary.append([step, tab_name, val_type, status, mismatches_count, matches_count, timestamp])
        
        # Format summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_summary['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        ws_summary['A7'].font = Font(bold=True)
        ws_summary['A7'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for col in range(1, 8):
            ws_summary[f'{get_column_letter(col)}7'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws_summary[f'{get_column_letter(col)}7'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ws_summary.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[col_letter].width = adjusted_width
        
        # ============================================================================
        # Sheet 2: All Mismatches
        # ============================================================================
        ws_mismatches = wb.create_sheet("All Mismatches")
        ws_mismatches.append(["Step", "Tab Name", "Row", "Column", "Expected", "Actual", "Match Type", "Node Type"])
        
        # Collect all mismatches
        for v in validations:
            step = v.get('step', '')
            tab_name = v.get('webTabName', '') or v.get('excelTabName', '')
            val_type = v.get('validationType', 'table')
            
            # Regular mismatches
            for m in v.get('mismatches', []):
                ws_mismatches.append([
                    step,
                    tab_name,
                    m.get('row', ''),
                    m.get('column', ''),
                    m.get('expected', ''),
                    m.get('actual', ''),
                    m.get('matchType', 'exact'),
                    ''  # Node type (only for data_view)
                ])
            
            # Node results mismatches (for data_view)
            if val_type == 'data_view':
                node_results = v.get('nodeResults', [])
                for nr in node_results:
                    node_type = nr.get('nodeType', '')
                    for m in nr.get('mismatches', []):
                        ws_mismatches.append([
                            step,
                            tab_name,
                            m.get('row', ''),
                            m.get('column', ''),
                            m.get('expected', ''),
                            m.get('actual', ''),
                            m.get('matchType', 'exact'),
                            node_type
                        ])
        
        # Format mismatches sheet header
        for col in range(1, 9):
            cell = ws_mismatches[f'{get_column_letter(col)}1']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for col in ws_mismatches.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_mismatches.column_dimensions[col_letter].width = adjusted_width
        
        # ============================================================================
        # Sheet 3: Node Results (for data_view validations)
        # ============================================================================
        data_view_validations = [v for v in validations if v.get('validationType') == 'data_view']
        if data_view_validations:
            ws_nodes = wb.create_sheet("Node Results")
            ws_nodes.append(["Step", "Tab Name", "Node Type", "Status", "Mismatch Count", "Error"])
            
            for v in data_view_validations:
                step = v.get('step', '')
                tab_name = v.get('webTabName', '') or v.get('excelTabName', '')
                node_results = v.get('nodeResults', [])
                
                for nr in node_results:
                    node_type = nr.get('nodeType', '')
                    success = nr.get('success', False)
                    mismatch_count = len(nr.get('mismatches', []))
                    error = nr.get('error', '')
                    status = "✅ Passed" if success else "❌ Failed"
                    
                    ws_nodes.append([step, tab_name, node_type, status, mismatch_count, error])
            
            # Format node results sheet header
            for col in range(1, 7):
                cell = ws_nodes[f'{get_column_letter(col)}1']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for col in ws_nodes.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_nodes.column_dimensions[col_letter].width = adjusted_width
        
        # ============================================================================
        # Save to BytesIO and return
        # ============================================================================
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f'{execution_id}_validation_results.xlsx'
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(f"Error generating Excel validation results: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/executions', methods=['GET'])
def list_executions():
    project_root = current_app.config['PROJECT_ROOT']
    results_dir = project_root / 'storage' / 'executions'
    executions = []
    
    if results_dir.exists():
        for f in results_dir.glob('*.json'):
            try:
                with open(f) as file:
                    r = json.load(file)
                
                # Format story text - handle Excel executions differently
                story_text = r.get('story', 'No test scenario')
                if r.get('source') == 'excel':
                    # For Excel executions, show Excel filename
                    excel_meta = r.get('excel_metadata', {})
                    excel_filename = excel_meta.get('filename', 'Excel Test')
                    story_text = f"📊 {excel_filename}"
                else:
                    # Truncate long stories
                    story_text = story_text[:100]
                
                executions.append({
                    'execution_id': r['execution_id'],
                    'story': story_text,
                    'status': r['status'],
                    'source': r.get('source', 'ai'),  # 'ai' or 'excel'
                    'actions_count': len(r.get('actions_taken', [])),
                    'screenshots_count': len(r.get('screenshots', [])),
                    'started_at': r.get('started_at') or r.get('created_at'),
                    'completed_at': r.get('completed_at'),
                    'duration': r.get('duration')
                })
            except:
                continue
        
        # Sort by started_at timestamp (most recent first)
        # Handle mixed types (string vs float) by converting to comparable format
        def get_sort_key(exec):
            started_at = exec.get('started_at') or exec.get('created_at') or ''
            if isinstance(started_at, (int, float)):
                return started_at
            elif isinstance(started_at, str):
                # Try to parse ISO format or return 0
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(started_at.replace('Z', '+00:00'))
                    return dt.timestamp()
                except:
                    return 0
            return 0
        
        executions.sort(key=get_sort_key, reverse=True)
    
    return jsonify({'executions': executions}), 200


@bp.route('/screenshots/<path:filename>', methods=['GET'])
def get_screenshot(filename):
    project_root = current_app.config['PROJECT_ROOT']
    
    # Try main screenshots directory first
    path = project_root / 'storage' / 'screenshots' / filename
    if path.exists():
        return send_file(path, mimetype='image/png')
    
    # Also check test directory screenshots (for Excel tests)
    test_path = project_root / 'storage' / 'excel_tests' / 'storage' / 'screenshots' / filename
    if test_path.exists():
        return send_file(test_path, mimetype='image/png')
    
    return jsonify({'error': 'Not found'}), 404


@bp.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'architecture': 'Pure Python + Playwright'}), 200


# Element Map Manager Routes
@bp.route('/fetch-html', methods=['POST'])
def fetch_html():
    """Fetch HTML from URL using Playwright"""
    try:
        data = request.json
        url = data.get('url', '')
        
        if not url:
            return jsonify({'error': 'URL is required'}), 400
        
        # Import Playwright fetcher
        import asyncio
        from playwright.async_api import async_playwright
        
        async def fetch_page_html(url):
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                await page.goto(url, wait_until='networkidle', timeout=30000)
                html = await page.content()
                await browser.close()
                return html
        
        html = asyncio.run(fetch_page_html(url))
        
        return jsonify({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route('/parse-html', methods=['POST'])
def parse_html():
    """Parse HTML using Playwright parser with live DOM (for XPath testing)"""
    try:
        data = request.json
        url = data.get('url', '')
        page_name = data.get('page_name', '')
        
        if not url:
            return jsonify({'error': 'URL is required'}), 400
        
        # Import Playwright parser
        import sys
        import asyncio
        from pathlib import Path
        sys.path.insert(0, str(Path(current_app.config['PROJECT_ROOT'])))
        from playwright.async_api import async_playwright
        from utils.playwright_tree_parser import parse_with_tree
        
        async def parse_with_playwright(url, page_name):
            """Parse page using Playwright with live DOM"""
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                # Set larger viewport to ensure all tabs and elements are visible
                page = await browser.new_page(viewport={'width': 1920, 'height': 1080})
                
                # Navigate to URL
                await page.goto(url, wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(3000)
                
                # Dismiss any popups
                try:
                    continue_btn = page.locator("text='Continue'").first
                    if await continue_btn.is_visible(timeout=2000):
                        await continue_btn.click()
                        await page.wait_for_timeout(1000)
                except:
                    pass
                
                # Parse using tree-based parser with live DOM
                element_map = await parse_with_tree(page)
                
                await browser.close()
                
                # Override page name if provided
                if page_name:
                    element_map["page"] = page_name
                
                return element_map
        
        # Run async parser
        element_map = asyncio.run(parse_with_playwright(url, page_name))
        
        return jsonify({
            'success': True,
            'element_map': element_map
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/parse-html-string', methods=['POST'])
def parse_html_string():
    """
    Parse HTML string and generate element registry JSON
    
    Expected JSON:
    {
        "html": "<html>...</html>",
        "url": "https://example.com/page",
        "page_name": "home"
    }
    
    Returns:
        Element registry JSON in the exact format
    """
    try:
        data = request.get_json()
        html_string = data.get('html', '').strip()
        url = data.get('url', '').strip()
        page_name = data.get('page_name', '').strip()
        
        if not html_string:
            return jsonify({'error': 'HTML content is required'}), 400
        if not url:
            return jsonify({'error': 'URL is required'}), 400
        if not page_name:
            return jsonify({'error': 'Page name is required'}), 400
        
        # Import parser
        try:
            from utils.html_registry_parser import HTMLRegistryParser
        except ImportError as e:
            return jsonify({
                'error': f'Failed to import parser: {str(e)}'
            }), 500
        
        try:
            parser = HTMLRegistryParser()
            registry = parser.parse_html(html_string, url, page_name)
            
            # Debug: Log how many elements were found
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Parsed HTML: found {len(registry.get('elements', {}))} elements")
            
            return jsonify({
                'success': True,
                'registry': registry
            }), 200
        except Exception as parse_error:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Parse error: {parse_error}")
            logger.error(traceback.format_exc())
            return jsonify({
                'error': f'Parsing failed: {str(parse_error)}',
                'traceback': traceback.format_exc()
            }), 500
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/parser/save', methods=['POST'])
def save_parser_registry():
    """
    Save a parser registry with a unique name (URL or PageName)
    
    Expected JSON:
    {
        "registry": {...},  # The full registry JSON
        "name": "home",     # Unique name (URL or PageName)
        "url": "https://example.com/page",  # Optional: for reference
        "page_name": "home"  # Optional: for reference
    }
    
    Returns:
        Success message with saved registry info
    """
    try:
        data = request.get_json()
        registry = data.get('registry')
        name = data.get('name', '').strip()
        url = data.get('url', '').strip()
        page_name = data.get('page_name', '').strip()
        
        if not registry:
            return jsonify({'error': 'Registry data is required'}), 400
        if not name:
            return jsonify({'error': 'Registry name is required'}), 400
        
        # Sanitize name (remove invalid characters for filename)
        import re
        sanitized_name = re.sub(r'[<>:"/\\|?*]', '_', name)
        sanitized_name = sanitized_name.strip()
        if not sanitized_name:
            return jsonify({'error': 'Invalid registry name. Please use alphanumeric characters and common symbols.'}), 400
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        parser_registries_dir = project_root / 'storage' / 'parser_registries'
        parser_registries_dir.mkdir(parents=True, exist_ok=True)
        
        # Check for duplicate name
        registry_file = parser_registries_dir / f"{sanitized_name}.json"
        if registry_file.exists():
            return jsonify({
                'error': f'Registry with name "{sanitized_name}" already exists. Please choose a different name.',
                'duplicate': True
            }), 400
        
        # Use sanitized name
        name = sanitized_name
        
        # Create registry metadata
        registry_metadata = {
            'name': name,
            'url': url,
            'page_name': page_name,
            'saved_at': datetime.now().isoformat(),
            'registry': registry
        }
        
        # Save registry
        with open(registry_file, 'w') as f:
            json.dump(registry_metadata, f, indent=2)
        
        return jsonify({
            'success': True,
            'message': f'Registry "{name}" saved successfully',
            'name': name,
            'saved_at': registry_metadata['saved_at']
        }), 200
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/parser/list', methods=['GET'])
def list_parser_registries():
    """
    List all saved parser registries
    
    Returns:
        List of saved registries with metadata
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        parser_registries_dir = project_root / 'storage' / 'parser_registries'
        
        if not parser_registries_dir.exists():
            return jsonify({
                'success': True,
                'registries': []
            }), 200
        
        registries = []
        for registry_file in parser_registries_dir.glob('*.json'):
            try:
                with open(registry_file, 'r') as f:
                    metadata = json.load(f)
                
                # Extract summary info
                registry_data = metadata.get('registry', {})
                registries.append({
                    'name': metadata.get('name', registry_file.stem),
                    'url': metadata.get('url', ''),
                    'page_name': metadata.get('page_name', ''),
                    'saved_at': metadata.get('saved_at', ''),
                    'element_count': len(registry_data.get('elements', {}))
                })
            except Exception as e:
                continue
        
        # Sort by saved_at (most recent first)
        registries.sort(key=lambda x: x.get('saved_at', ''), reverse=True)
        
        return jsonify({
            'success': True,
            'registries': registries
        }), 200
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/parser/load/<name>', methods=['GET'])
def load_parser_registry(name):
    """
    Load a saved parser registry by name
    
    Returns:
        Full registry JSON
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        parser_registries_dir = project_root / 'storage' / 'parser_registries'
        registry_file = parser_registries_dir / f"{name}.json"
        
        if not registry_file.exists():
            return jsonify({
                'error': f'Registry "{name}" not found'
            }), 404
        
        with open(registry_file, 'r') as f:
            metadata = json.load(f)
        
        return jsonify({
            'success': True,
            'registry': metadata.get('registry'),
            'name': metadata.get('name'),
            'url': metadata.get('url', ''),
            'page_name': metadata.get('page_name', ''),
            'saved_at': metadata.get('saved_at', '')
        }), 200
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/parser/delete/<name>', methods=['DELETE'])
def delete_parser_registry(name):
    """
    Delete a saved parser registry by name
    
    Returns:
        Success message
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        parser_registries_dir = project_root / 'storage' / 'parser_registries'
        registry_file = parser_registries_dir / f"{name}.json"
        
        if not registry_file.exists():
            return jsonify({
                'error': f'Registry "{name}" not found'
            }), 404
        
        registry_file.unlink()
        
        return jsonify({
            'success': True,
            'message': f'Registry "{name}" deleted successfully'
        }), 200
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/manual-register', methods=['POST'])
def manual_register_element():
    """
    Manually register element from HTML
    User pastes HTML element + URL, system updates registry
    
    Expected JSON:
    {
        "html": "<input name=\"name\" data-testid=\"submission-name-input\"...>",
        "url": "https://hub-stage.datacommons.cancer.gov/data-submissions",
        "element_name": "Submission Name"  // Optional - will infer if not provided
    }
    """
    try:
        data = request.get_json()
        html_string = data.get('html', '').strip()
        url = data.get('url', '').strip()
        element_name = data.get('element_name', '').strip() or None
        
        # Validation
        if not html_string:
            return jsonify({'success': False, 'error': 'HTML element required'}), 400
        if not url:
            return jsonify({'success': False, 'error': 'URL required'}), 400
        
        # Initialize registry
        element_registry = get_registry()
        
        # Register element
        from api.manual_registry_helper import ManualRegistryHelper
        helper = ManualRegistryHelper(element_registry)
        
        # Debug: Log received HTML (first 200 chars)
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Manual register - HTML received (first 200 chars): {html_string[:200]}")
        logger.info(f"Manual register - URL: {url}, Element name: {element_name}")
        
        result = helper.register_element(
            html_string=html_string,
            url=url,
            element_name=element_name
        )
        
        # Debug: Log result
        logger.info(f"Manual register - Result: success={result.get('success')}, uniqueness_method={result.get('uniqueness_method')}, attributes_found={result.get('attributes_found', [])}")
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify(result), 400
            
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'details': traceback.format_exc()
        }), 500


@bp.route('/save-element-map', methods=['POST'])
def save_element_map():
    """Save parsed element map to registry"""
    try:
        data = request.json
        element_map = data.get('element_map')
        
        if not element_map:
            return jsonify({'error': 'Element map is required'}), 400
        
        # Import registry
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(current_app.config['PROJECT_ROOT'])))
        from utils.element_registry import get_registry
        
        registry = get_registry(str(Path(current_app.config['PROJECT_ROOT']) / 'element_maps'))
        
        # Extract domain and page from URL
        url = element_map.get('url', '')
        domain = url.replace('https://', '').replace('http://', '').split('/')[0].split('#')[0]
        page = element_map.get('page', 'unknown')
        
        # Sanitize page name (remove URL components if present)
        if page.startswith('http://') or page.startswith('https://'):
            # Page is a full URL, extract the page name from it
            page_clean = page.replace('https://', '').replace('http://', '')
            page_clean = page_clean.split('/')[0].split('#')[0]  # Remove domain
            # Extract last path segment or use 'home' if root
            page_parts = page.split('/')[-1].split('#')
            if page_parts and page_parts[0]:
                page = page_parts[0]
            else:
                page = 'home'
        
        # Remove any invalid filesystem characters from page name
        page = page.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
        
        # If page is still empty or too long, use default
        if not page or len(page) > 100:
            page = 'home'
        
        # Update element_map with cleaned page name
        element_map['page'] = page
        
        # Save to registry
        registry.save_map(domain, page, element_map)
        
        # Create baseline
        registry.create_baseline(domain, page)
        
        map_path = registry.get_map_path(domain, page)
        
        return jsonify({
            'success': True,
            'message': f'Element map saved successfully',
            'path': str(map_path),
            'domain': domain,
            'page': page
        })
        
    except Exception as e:
        import traceback
        import logging
        error_details = traceback.format_exc()
        logger = logging.getLogger(__name__)
        logger.error(f"❌ ERROR in save_element_map: {e}")
        logger.error(f"Full traceback:\n{error_details}")
        current_app.logger.error(f"❌ ERROR in save_element_map: {e}")
        current_app.logger.error(f"Full traceback:\n{error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

@bp.route('/element-maps/list')
def list_element_maps():
    """List all existing element maps"""
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(current_app.config['PROJECT_ROOT'])))
        
        maps_dir = Path(current_app.config['PROJECT_ROOT']) / 'element_maps'
        
        maps = []
        for domain_dir in maps_dir.iterdir():
            if domain_dir.is_dir() and domain_dir.name not in ['versions', '__pycache__']:
                domain = domain_dir.name
                for map_file in domain_dir.glob('*_page.json'):
                    if map_file.is_file():
                        import json
                        with open(map_file, 'r') as f:
                            map_data = json.load(f)
                        
                        maps.append({
                            'domain': domain,
                            'page': map_data.get('page'),
                            'url': map_data.get('url'),
                            'version': map_data.get('version'),
                            'total_elements': map_data.get('statistics', {}).get('total_elements', 0),
                            'last_updated': map_data.get('last_updated'),
                            'file': str(map_file)
                        })
        
        return jsonify({'maps': maps})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route('/element-maps/<domain>/<page>')
def get_element_map(domain, page):
    """Get specific element map"""
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(current_app.config['PROJECT_ROOT'])))
        from utils.element_registry import get_registry
        
        registry = get_registry(str(Path(current_app.config['PROJECT_ROOT']) / 'element_maps'))
        element_map = registry.load_map(domain, page)
        
        if not element_map:
            return jsonify({'error': 'Map not found'}), 404
        
        return jsonify(element_map)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<execution_id>/approve-discoveries', methods=['POST'])
def approve_discoveries(execution_id):
    """
    User approved test execution - update element registry with discoveries
    
    This commits the discovered selectors to the registry so future tests
    can use the optimized selectors instead of repeating discovery.
    """
    try:
        import sys
        from pathlib import Path
        from urllib.parse import urlparse
        sys.path.insert(0, str(Path(current_app.config['PROJECT_ROOT'])))
        from utils.element_registry import get_registry
        
        project_root = current_app.config['PROJECT_ROOT']
        
        # Load discoveries from file
        discoveries_dir = project_root / 'storage' / 'discoveries'
        discovery_file = discoveries_dir / f'{execution_id}_discoveries.json'
        
        if not discovery_file.exists():
            return jsonify({
                'error': 'Discovery file not found',
                'execution_id': execution_id
            }), 404
        
        with open(discovery_file, 'r') as f:
            discovery_data = json.load(f)
        
        discoveries = discovery_data.get('discoveries', [])
        
        if not discoveries:
            return jsonify({
                'error': 'No discoveries found in this execution',
                'execution_id': execution_id
            }), 400
        
        # Get registry
        registry = get_registry(str(project_root / 'element_maps'))
        
        # Extract domain from execution results
        results_file = project_root / 'storage' / 'executions' / f'{execution_id}.json'
        
        if not results_file.exists():
            return jsonify({'error': 'Execution results not found'}), 404
        
        with open(results_file, 'r') as f:
            results = json.load(f)
        
        # Get domain from story or first action
        story = results.get('story', '')
        domain = None
        
        # Try to extract domain from story URL
        if 'https://' in story or 'http://' in story:
            import re
            url_match = re.search(r'https?://([^\s/]+)', story)
            if url_match:
                domain = url_match.group(1)
        
        if not domain:
            return jsonify({'error': 'Could not determine domain from test execution'}), 400
        
        page = "home"  # Default page name
        
        # Update registry with each discovery
        updated_count = 0
        for discovery in discoveries:
            try:
                registry.update_with_discovery(domain, page, discovery)
                updated_count += 1
            except Exception as e:
                print(f"Warning: Failed to update discovery {discovery.get('name')}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'message': f'Registry updated with {updated_count} discoveries',
            'execution_id': execution_id,
            'discoveries_updated': updated_count,
            'domain': domain,
            'page': page
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error approving discoveries: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/generate-and-validate', methods=['POST'])
def generate_and_validate(exec_id):
    """Generate Playwright test and validate it"""
    try:
        project_root = current_app.config['PROJECT_ROOT']
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        
        # Check if this is an Excel execution
        if execution_file.exists():
            with open(execution_file, 'r') as f:
                exec_data = json.load(f)
            
            if exec_data.get('source') == 'excel':
                # For Excel executions, regenerate from Excel file
                excel_id = exec_data.get('excel_id')
                if not excel_id:
                    return jsonify({
                        'error': 'Excel ID not found in execution data',
                        'execution_id': exec_id
                    }), 400
                
                # Import TypeScript Excel generation function
                sys.path.insert(0, str(project_root))
                from REFACTOR.generator.excel_generator_ts import generate_playwright_ts_from_excel
                from validator.typescript_test_runner import TypeScriptTestRunner
                from pathlib import Path
                from datetime import datetime
                import uuid
                
                # Get Excel file path from metadata
                metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
                metadata_file = metadata_dir / f"{excel_id}.json"
                
                if not metadata_file.exists():
                    return jsonify({
                        'error': f'Excel metadata not found: {excel_id}'
                    }), 404
                
                with open(metadata_file, 'r') as f:
                    metadata = json.load(f)
                
                excel_path = project_root / metadata['file_path']
                if not excel_path.exists():
                    return jsonify({
                        'error': f'Excel file not found: {excel_path}'
                    }), 404
                
                # Generate new TypeScript test file
                output_dir = project_root / 'storage' / 'excel_tests'
                output_dir.mkdir(parents=True, exist_ok=True)
                new_exec_id = f"excel_exec_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
                output_file = output_dir / f"test_excel_{excel_id}.spec.ts"
                
                # Generate TypeScript test from Excel
                generation_result = generate_playwright_ts_from_excel(excel_path, output_file)
                
                if not generation_result.get('success'):
                    return jsonify({
                        'error': 'Failed to generate test from Excel',
                        'errors': generation_result.get('errors', [])
                    }), 500
                
                # Update execution data with new test file
                exec_data['test_file'] = str(output_file.relative_to(project_root))
                exec_data['regenerated_at'] = datetime.now().isoformat()
                with open(execution_file, 'w') as f:
                    json.dump(exec_data, f, indent=2)
                
                # Run TypeScript test in background
                def run_test_background():
                    try:
                        runner = TypeScriptTestRunner(project_root)
                        test_result = runner.run(str(output_file), exec_id)
                        
                        # Update execution with test results
                        exec_data['playwright_validation'] = {
                            'status': test_result.get('status'),
                            'duration': test_result.get('duration'),
                            'assertions_passed': test_result.get('assertions_passed', 0),
                            'assertions_failed': test_result.get('assertions_failed', 0),
                            'test_file': test_result.get('test_file'),
                            'timestamp': test_result.get('timestamp'),
                            'stdout': test_result.get('stdout', ''),
                            'stderr': test_result.get('stderr', ''),
                            'exit_code': test_result.get('exit_code', 0),
                            'validation_mismatches': test_result.get('validation_mismatches', [])
                        }
                        exec_data['playwright_screenshots'] = test_result.get('screenshots', [])
                        # Update execution status based on test result
                        test_status = test_result.get('status', 'unknown')
                        if test_status == 'failed' or test_result.get('exit_code', 0) != 0:
                            exec_data['status'] = 'failed'
                        else:
                            exec_data['status'] = 'completed'
                        exec_data['completed_at'] = datetime.now().isoformat()
                        
                        with open(execution_file, 'w') as f:
                            json.dump(exec_data, f, indent=2)
                    except Exception as e:
                        print(f"Error running Excel test in background: {e}")
                        import traceback
                        traceback.print_exc()
                
                thread = threading.Thread(target=run_test_background)
                thread.daemon = True
                thread.start()
                
                return jsonify({
                    'success': True,
                    'execution_id': exec_id,
                    'test_file': str(output_file.relative_to(project_root)),
                    'test_path': str(output_file),
                    'regenerated_at': exec_data['regenerated_at'],
                    'test_running': True,
                    'source': 'excel'
                }), 200
        
        # Regular execution flow (original code)
        # Force reload to ensure latest code (reload all generator modules)
        import importlib
        modules_to_reload = [
            'generator.playwright_generator',
            'generator.pw_codegen.step_generators',
            'generator.pw_core.generator',
        ]
        for module_name in modules_to_reload:
            if module_name in sys.modules:
                importlib.reload(sys.modules[module_name])
        from generator.playwright_generator import PlaywrightGenerator
        from validator.test_runner import TestRunner
        from validator.comparator import Comparator
        
        # Handle empty body gracefully
        data = request.get_json(silent=True) or {}
        test_name = data.get('test_name')
        validate_selectors = data.get('validate_selectors', True)  # Pre-generation selector validation
        
        # Step 1: Generate Playwright code (with pre-validation)
        try:
            generator = PlaywrightGenerator(project_root)
            generation_result = generator.generate(exec_id, test_name, validate_selectors=validate_selectors)
        except Exception as e:
            # If selector validation fails, return detailed error
            error_msg = str(e)
            if "Selector Validation Failed" in error_msg:
                return jsonify({
                    'success': False,
                    'error': error_msg,
                    'error_type': 'selector_validation_failed',
                    'execution_id': exec_id
                }), 400
            # Re-raise other errors
            raise
        
        result = {
            'success': True,
            'execution_id': exec_id,
            'test_file': generation_result['filename'],
            'test_path': generation_result['filepath'],
            'code_preview': generation_result['code'][:500] + '...',  # Preview only
            'generated_at': generation_result['metadata']['generated_at'],
            'test_running': True  # Indicate test is running
        }
        
        # Step 2: Run test in background thread (automatic, no separate button needed)
        def run_test_background():
            try:
                runner = TestRunner(project_root)
                test_result = runner.run(generation_result['filename'], exec_id)
                
                # Step 3: Compare results
                comparator = Comparator(project_root)
                comparison = comparator.compare(exec_id, test_result)
                
                # Step 4: Save Playwright results back to execution file for UI display
                results_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
                if results_file.exists():
                    with open(results_file, 'r') as f:
                        exec_data = json.load(f)
                    
                    # Add Playwright test results to execution data
                    exec_data['playwright_screenshots'] = test_result.get('screenshots', [])
                    exec_data['playwright_validation'] = {
                        'status': test_result.get('status'),
                        'duration': test_result.get('duration'),
                        'assertions_passed': test_result.get('assertions_passed'),
                        'assertions_failed': test_result.get('assertions_failed'),
                        'test_file': test_result.get('test_file'),
                        'timestamp': test_result.get('timestamp'),
                        'stdout': test_result.get('stdout', ''),
                        'stderr': test_result.get('stderr', ''),
                        'exit_code': test_result.get('exit_code', 0),
                        'validation_mismatches': test_result.get('validation_mismatches', [])
                    }
                    exec_data['playwright_comparison'] = comparison
                    
                    # Write back to file
                    with open(results_file, 'w') as f:
                        json.dump(exec_data, f, indent=2)
            except Exception as e:
                print(f"Error running test in background: {e}")
                import traceback
                traceback.print_exc()
        
        # Start background thread to run test
        thread = threading.Thread(target=run_test_background)
        thread.daemon = True
        thread.start()
        
        return jsonify(result), 200
        
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        import traceback
        print(f"Error generating/validating test: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/generated-test', methods=['GET'])
def get_generated_test(exec_id):
    """Get generated test code - supports both regular and Excel executions"""
    try:
        project_root = current_app.config['PROJECT_ROOT']
        
        # Check if this is an Excel execution
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        if execution_file.exists():
            with open(execution_file, 'r') as f:
                exec_data = json.load(f)
            
            # Excel execution has test_file directly in execution data
            if exec_data.get('source') == 'excel' and exec_data.get('test_file'):
                test_file_path = project_root / exec_data['test_file']
                if test_file_path.exists():
                    with open(test_file_path, 'r') as f:
                        code = f.read()
                    return jsonify({
                        'execution_id': exec_id,
                        'test_file': exec_data['test_file'],
                        'test_name': exec_data.get('test_name', 'excel_test'),
                        'code': code,
                        'source': 'excel'
                    }), 200
        
        # Fallback to regular generated test metadata
        metadata_file = project_root / 'storage' / 'generated_tests' / f'{exec_id}_test.json'
        
        if not metadata_file.exists():
            return jsonify({'error': 'No generated test found for this execution'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Read the generated code
        test_file = Path(metadata['filename'])
        if test_file.exists():
            with open(test_file, 'r') as f:
                code = f.read()
            metadata['code'] = code
        
        return jsonify(metadata), 200
        
    except Exception as e:
        print(f"Error retrieving generated test: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/run-test', methods=['POST'])
def run_test(exec_id):
    """Run generated TypeScript Playwright test in background to generate screenshots"""
    try:
        from validator.typescript_test_runner import TypeScriptTestRunner
        from datetime import datetime
        
        project_root = current_app.config['PROJECT_ROOT']
        
        # Check execution file for test file path
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        if not execution_file.exists():
            return jsonify({'error': 'Execution not found'}), 404
        
        with open(execution_file, 'r') as f:
            exec_data = json.load(f)
        
        # Get test file from execution data (should be TypeScript .spec.ts file)
        test_file_path = exec_data.get('test_file')
        if not test_file_path:
            return jsonify({'error': 'Test file not found in execution data'}), 404
        
        # Resolve full path
        test_path = project_root / test_file_path
        if not test_path.exists():
            return jsonify({'error': f'Test file not found: {test_path}'}), 404
        
        # Run TypeScript test in background thread to avoid timeout
        def run_test_background():
            try:
                runner = TypeScriptTestRunner(project_root)
                test_result = runner.run(str(test_path), exec_id)
                
                # Save results to execution file
                results_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
                if results_file.exists():
                    with open(results_file, 'r') as f:
                        exec_data = json.load(f)
                    
                    exec_data['playwright_screenshots'] = test_result.get('screenshots', [])
                    exec_data['playwright_validation'] = {
                        'status': test_result.get('status'),
                        'duration': test_result.get('duration'),
                        'assertions_passed': test_result.get('assertions_passed', 0),
                        'assertions_failed': test_result.get('assertions_failed', 0),
                        'test_file': test_result.get('test_file'),
                        'timestamp': test_result.get('timestamp'),
                        'stdout': test_result.get('stdout', ''),
                        'stderr': test_result.get('stderr', ''),
                        'exit_code': test_result.get('exit_code', 0)
                    }
                    # Update execution status based on test result
                    test_status = test_result.get('status', 'unknown')
                    if test_status == 'failed' or test_result.get('exit_code', 0) != 0:
                        exec_data['status'] = 'failed'
                    else:
                        exec_data['status'] = 'completed'
                    exec_data['completed_at'] = datetime.now().isoformat()
                    
                    with open(results_file, 'w') as f:
                        json.dump(exec_data, f, indent=2)
            except Exception as e:
                print(f"Error running test in background: {e}")
                import traceback
                traceback.print_exc()
        
        # Start background thread
        thread = threading.Thread(target=run_test_background)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Test execution started in background. Screenshots will be available shortly.',
            'execution_id': exec_id
        }), 202  # 202 Accepted (async processing)
        
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 404
    except Exception as e:
        import traceback
        print(f"Error starting test execution: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/download-test', methods=['GET'])
def download_generated_test(exec_id):
    """Download generated test code as .py file for standalone execution - supports both regular and Excel executions"""
    try:
        from flask import send_file
        project_root = current_app.config['PROJECT_ROOT']
        
        # Check if this is an Excel execution
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        if execution_file.exists():
            with open(execution_file, 'r') as f:
                exec_data = json.load(f)
            
            # Excel execution has test_file directly in execution data
            if exec_data.get('source') == 'excel' and exec_data.get('test_file'):
                test_file_path = project_root / exec_data['test_file']
                if test_file_path.exists():
                    filename = exec_data.get('test_name', 'excel_test') + '.py'
                    response = send_file(
                        str(test_file_path),
                        mimetype='text/x-python',
                        as_attachment=True,
                        download_name=filename,
                    )
                    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                    response.headers['Pragma'] = 'no-cache'
                    response.headers['Expires'] = '0'
                    return response
                else:
                    # Excel execution but test file doesn't exist
                    return jsonify({
                        'error': f'Excel test file not found: {test_file_path}',
                        'test_file': exec_data.get('test_file'),
                        'resolved_path': str(test_file_path.resolve())
                    }), 404
        
        # Fallback to regular generated test metadata
        metadata_file = project_root / 'storage' / 'generated_tests' / f'{exec_id}_test.json'
        
        if not metadata_file.exists():
            return jsonify({'error': 'No generated test found for this execution'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get the test file path
        test_file = Path(metadata['filename'])
        if not test_file.exists():
            return jsonify({'error': 'Generated test file not found'}), 404
        
        # Extract just the filename for download
        filename = test_file.name
        
        # Send file with proper headers for download
        # Convert Path to string for Flask compatibility
        response = send_file(
            str(test_file),
            as_attachment=True,
            download_name=filename,
            mimetype='text/x-python'
        )
        # Prevent browser caching - force fresh download every time
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
        
    except Exception as e:
        print(f"Error downloading generated test: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/download-test-zip', methods=['GET'])
def download_test_with_env_zip(exec_id):
    """Download test file and all required registry JSON files bundled as a zip file (excludes .env for security) - supports both regular and Excel executions"""
    try:
        from flask import send_file
        from io import BytesIO
        import zipfile
        import os
        import re
        
        project_root = current_app.config['PROJECT_ROOT']
        test_file = None
        test_name = None
        
        # Check if this is an Excel execution
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        if execution_file.exists():
            with open(execution_file, 'r') as f:
                exec_data = json.load(f)
            
            # Excel execution has test_file directly in execution data
            if exec_data.get('source') == 'excel' and exec_data.get('test_file'):
                test_file = project_root / exec_data['test_file']
                test_name = exec_data.get('test_name', 'excel_test')
                
                if not test_file.exists():
                    return jsonify({
                        'error': f'Excel test file not found: {test_file}',
                        'test_file': exec_data.get('test_file'),
                        'resolved_path': str(test_file.resolve())
                    }), 404
        
        # Fallback to regular generated test metadata
        if test_file is None:
            metadata_file = project_root / 'storage' / 'generated_tests' / f'{exec_id}_test.json'
            
            if not metadata_file.exists():
                return jsonify({'error': 'No generated test found for this execution'}), 404
            
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
            
            # Get the test file path
            test_file = Path(metadata['filename'])
            test_name = test_file.stem
            
            if not test_file.exists():
                return jsonify({'error': 'Generated test file not found'}), 404
        
        # Read test file to extract REGISTRY_PATHS
        with open(test_file, 'r') as f:
            test_content = f.read()
        
        # Extract REGISTRY_PATHS from test file using regex
        registry_paths = []
        # Match: REGISTRY_PATHS = [\n    'path1',\n    'path2',\n]
        match = re.search(r"REGISTRY_PATHS\s*=\s*\[(.*?)\]", test_content, re.DOTALL)
        if match:
            paths_str = match.group(1)
            # Extract all quoted strings
            path_matches = re.findall(r"['\"]([^'\"]+)['\"]", paths_str)
            registry_paths = [p.strip() for p in path_matches if p.strip()]
        
        # Create zip file in memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add test file
            zip_file.write(test_file, test_file.name)
            print(f"✅ Added test file to zip: {test_file.name}")
            
            # Note: .env file is NOT included for security reasons
            # Users must create their own .env file with appropriate credentials
            
            # Add all registry JSON files with proper directory structure
            # The test expects paths like 'element_maps/domain/page_page.json'
            # So we preserve that structure in the zip
            element_maps_dir = project_root / 'element_maps'
            for registry_path in registry_paths:
                # registry_path is like 'element_maps/domain/page_page.json'
                full_registry_path = project_root / registry_path
                
                if full_registry_path.exists():
                    # Preserve the full path structure as the test expects it
                    # e.g., 'element_maps/domain/page_page.json' stays as 'element_maps/domain/page_page.json'
                    zip_path = registry_path
                    
                    zip_file.write(full_registry_path, zip_path)
                    print(f"✅ Added registry file to zip: {zip_path}")
                else:
                    print(f"⚠️  Registry file not found: {full_registry_path}")
        
        # Verify zip contents
        zip_buffer.seek(0)
        with zipfile.ZipFile(zip_buffer, 'r') as verify_zip:
            file_list = verify_zip.namelist()
            print(f"📦 Zip contains {len(file_list)} files:")
            for name in file_list:
                print(f"   - {name}")
        
        zip_buffer.seek(0)
        
        # Extract test filename for zip name
        test_filename = test_file.name.replace('.py', '')
        zip_filename = f'{test_filename}_complete.zip'
        
        # Send zip file
        response = send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
        # Set headers
        response.headers['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"Error creating zip file: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/mark-passed', methods=['POST'])
def mark_test_passed(exec_id):
    """
    Mark test execution as passed and populate registry with Excel XPaths
    This allows future tests to use registry IDs instead of hard-coded XPaths
    """
    try:
        project_root = current_app.config['PROJECT_ROOT']
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        
        if not execution_file.exists():
            return jsonify({'error': 'Execution not found'}), 404
        
        # Load execution data
        with open(execution_file, 'r') as f:
            exec_data = json.load(f)
        
        # Check if this is an Excel execution
        if exec_data.get('source') != 'excel':
            return jsonify({
                'error': 'Only Excel executions can be marked as passed for registry update',
                'source': exec_data.get('source')
            }), 400
        
        # Get Excel file path from excel_id
        excel_id = exec_data.get('excel_id')
        if not excel_id:
            return jsonify({'error': 'Excel ID not found in execution data'}), 400
        
        # Load Excel metadata to get file path
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({
                'error': f'Excel metadata not found: {excel_id}',
                'metadata_file': str(metadata_file)
            }), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get Excel file path from metadata
        excel_file_path = metadata.get('file_path')
        if not excel_file_path:
            return jsonify({
                'error': 'Excel file path not found in metadata',
                'excel_id': excel_id
            }), 400
        
        excel_file = project_root / excel_file_path
        if not excel_file.exists():
            return jsonify({
                'error': 'Excel file not found',
                'path': str(excel_file),
                'excel_id': excel_id
            }), 404
        
        # Registry is not updated from Excel - only lookup existing XPaths
        # All XPaths must already exist in the registry JSON file
        
        # Mark execution as passed
        exec_data['test_status'] = 'passed'
        exec_data['marked_passed_at'] = datetime.now().isoformat()
        
        # Save updated execution data
        with open(execution_file, 'w') as f:
            json.dump(exec_data, f, indent=2)
        
        return jsonify({
            'success': True,
            'message': 'Test marked as passed successfully',
            'execution_id': exec_id,
            'excel_file': str(excel_file),
            'marked_passed_at': exec_data['marked_passed_at']
        })
        
    except Exception as e:
        import traceback
        print(f"Error marking test as passed: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/executions/<exec_id>/download-test-ts-zip', methods=['GET'])
def download_test_ts_zip(exec_id):
    """Download TypeScript test file (.spec.ts), package.json, README, and all required registry JSON files bundled as a zip file (excludes .env for security)"""
    try:
        from flask import send_file
        from io import BytesIO
        import zipfile
        import os
        import re
        
        project_root = current_app.config['PROJECT_ROOT']
        
        # Check if this is an Excel execution
        execution_file = project_root / 'storage' / 'executions' / f'{exec_id}.json'
        ts_test_file = None
        test_name = None
        exec_data = None
        
        if execution_file.exists():
            with open(execution_file, 'r') as f:
                exec_data = json.load(f)
            
            # Excel execution - check for pre-generated TypeScript file first
            if exec_data.get('source') == 'excel':
                excel_id = exec_data.get('excel_id')
                if excel_id:
                    # Check Excel metadata for pre-generated TypeScript file
                    metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
                    metadata_file = metadata_dir / f"{excel_id}.json"
                    if metadata_file.exists():
                        with open(metadata_file, 'r') as f:
                            excel_metadata = json.load(f)
                        if 'generated_test_ts' in excel_metadata:
                            ts_test_file = project_root / excel_metadata['generated_test_ts']['test_file']
                            test_name = excel_metadata['generated_test_ts']['test_name']
                            if ts_test_file.exists():
                                # Use pre-generated TypeScript file (preferred)
                                pass
                            else:
                                ts_test_file = None
                
                # Excel executions must have TypeScript test file
                # Fallback to execution JSON's test_file if Excel metadata doesn't have it
                if ts_test_file is None:
                    if exec_data.get('test_file'):
                        ts_test_file = project_root / exec_data['test_file']
                        test_name = exec_data.get('test_name') or exec_id  # Use execution ID as fallback
                        if not ts_test_file.exists():
                            ts_test_file = None
                
                if ts_test_file is None:
                    return jsonify({
                        'error': 'No TypeScript test file found for Excel execution. Please regenerate the test.',
                        'excel_id': excel_id,
                        'execution_id': exec_id
                    }), 404
        
        # Check if TypeScript file is available (required - no Python conversion)
        if ts_test_file is None or not ts_test_file.exists():
            return jsonify({
                'error': 'No TypeScript test file found. Only TypeScript tests are supported.',
                'execution_id': exec_id
            }), 404
        
        # Initialize variables
        registry_paths = []
        test_file_paths = []
        ts_code = None
        
        # Use pre-generated TypeScript file (required - no Python conversion)
        if ts_test_file and ts_test_file.exists():
            # Read pre-generated TypeScript file
            with open(ts_test_file, 'r') as f:
                ts_code = f.read()
            
            # Extract REGISTRY_PATHS from TypeScript file
            match = re.search(r"const\s+REGISTRY_PATHS\s*=\s*\[(.*?)\]", ts_code, re.DOTALL)
            if match:
                paths_str = match.group(1)
                path_matches = re.findall(r"['\"]([^'\"]+)['\"]", paths_str)
                registry_paths = [p.strip() for p in path_matches if p.strip()]
            
            # Extract file upload paths from TypeScript file
            # Look for patterns like: path.resolve(projectRoot, 'storage/test_files/filename.tsv')
            file_path_pattern = r"path\.resolve\(projectRoot,\s*['\"]([^'\"]+)['\"]\)"
            file_path_matches = re.findall(file_path_pattern, ts_code)
            for file_path in file_path_matches:
                file_path = file_path.strip()
                if file_path and file_path not in test_file_paths:
                    test_file_paths.append(file_path)
        else:
            return jsonify({
                'error': 'No TypeScript test file found. Only TypeScript tests are supported.',
                'execution_id': exec_id
            }), 404
        
        # Create package.json content (with xlsx for Excel reading)
        package_json_content = '''{
  "name": "playwright-test",
  "version": "1.0.0",
  "scripts": {
    "test": "playwright test"
  },
  "dependencies": {
    "@playwright/test": "^1.40.0",
    "dotenv": "^16.0.0",
    "xlsx": "^0.18.5"
  }
}'''
        
        # Create README content (test_name already set above)
        if test_name is None:
            test_name = exec_id  # Use execution ID as fallback
        
        test_files_note = ""
        if test_file_paths:
            test_files_note = f"\n- `storage/test_files/` - Test data files for file upload steps ({len(test_file_paths)} file(s))"
        
        readme_content = f'''# {test_name} - Playwright TypeScript Test

## Setup Instructions

1. Extract this zip file
2. The `.env` file is already included with credentials
3. Run: `npm install`
4. Run: `npx playwright install chromium`
5. Run: `npx playwright test {test_name}.spec.ts --headed`

## Files Included
- `{test_name}.spec.ts` - Main test file
- `package.json` - Dependencies
- `generate_totp.py` - Python script for TOTP generation (called from TypeScript)
- `element_maps/` - JSON registry files with element XPath mappings{test_files_note}

## Important
- **`.env` file is included** with the package (contains TOTP secrets and credentials)
- Registry JSON files are included in the `element_maps/` directory structure
- Test data files are included in the `storage/test_files/` directory structure (if file uploads are used)
- The test uses `pyotp` via Python script for TOTP generation to ensure consistency with Python tests

## Notes
- Screenshots saved to `storage/screenshots/`
- Test uses registry-based element lookup (no hard-coded XPaths)
- File paths are resolved relative to the test file location (goes up 2 levels to project root)
'''
        
        # Create zip file in memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add TypeScript test file
            ts_filename = f'{test_name}.spec.ts'
            zip_file.writestr(ts_filename, ts_code)
            print(f"✅ Added TypeScript test file to zip: {ts_filename}")
            
            # Note: .env file is NOT included for security reasons
            # Users must create their own .env file with appropriate credentials
            
            # Add package.json
            zip_file.writestr('package.json', package_json_content)
            print(f"✅ Added package.json to zip")
            
            # Add README.md
            zip_file.writestr('README.md', readme_content)
            print(f"✅ Added README.md to zip")
            
            # Add generate_totp.py script (required for TOTP generation in TypeScript tests)
            totp_script_path = project_root / 'Test' / 'generate_totp.py'
            if totp_script_path.exists():
                zip_file.write(totp_script_path, 'generate_totp.py')
                print(f"✅ Added generate_totp.py to zip")
            else:
                print(f"⚠️  generate_totp.py not found at {totp_script_path}")
            
            # Add .env file from Test directory
            env_file_path = project_root / 'Test' / '.env'
            if env_file_path.exists():
                zip_file.write(env_file_path, '.env')
                print(f"✅ Added .env file")
            else:
                print(f"⚠️  .env file not found at {env_file_path}")
            
            # Add all registry JSON files
            for registry_path in registry_paths:
                full_registry_path = project_root / registry_path
                if full_registry_path.exists():
                    zip_path = registry_path
                    zip_file.write(full_registry_path, zip_path)
                    print(f"✅ Added registry file to zip: {zip_path}")
                else:
                    print(f"⚠️  Registry file not found: {full_registry_path}")
            
            # Add test files for file upload steps
            if test_file_paths:
                for file_path in test_file_paths:
                    # file_path is like 'storage/test_files/dcpagain-icdc_file.tsv'
                    full_file_path = project_root / file_path
                    
                    if full_file_path.exists():
                        # Preserve the full path structure as the test expects it
                        zip_path = file_path
                        zip_file.write(full_file_path, zip_path)
                        print(f"✅ Added test file to zip: {zip_path}")
                    else:
                        print(f"⚠️  Test file not found: {full_file_path}")
            else:
                print("ℹ️  No file upload paths found in test file")
            
            # Add Excel file for Excel executions (source file for credentials and expected results)
            if exec_data and exec_data.get('source') == 'excel':
                excel_id = exec_data.get('excel_id')
                if excel_id:
                    metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
                    metadata_file = metadata_dir / f"{excel_id}.json"
                    if metadata_file.exists():
                        with open(metadata_file, 'r') as f:
                            excel_metadata = json.load(f)
                        excel_path = project_root / excel_metadata['file_path']
                        excel_filename = excel_metadata.get('filename', f'{excel_id}.xlsx')
                        if excel_path.exists():
                            zip_file.write(excel_path, excel_filename)
                            print(f"✅ Added Excel file to zip: {excel_filename}")
                        else:
                            print(f"⚠️  Excel file not found: {excel_path}")
        
        zip_buffer.seek(0)
        
        # Create zip filename - use execution ID for execution-based downloads
        # This ensures filename matches the execution ID shown in UI
        if exec_data and exec_data.get('source') == 'excel':
            # For Excel executions, use execution ID in filename to match UI
            zip_filename = f'{exec_id}_typescript_complete.zip'
        else:
            # Fallback to test_name for non-Excel executions
            zip_filename = f'{test_name}_typescript_complete.zip'
        
        # Send zip file
        response = send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"Error creating TypeScript zip file: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/registry/<domain>/<page>/download', methods=['GET'])
def download_registry_json(domain, page):
    """Download registry JSON file"""
    try:
        from flask import send_file
        project_root = current_app.config['PROJECT_ROOT']
        
        # Try both naming conventions: {page}_page.json and {page}.json
        registry_file = project_root / 'element_maps' / domain / f'{page}_page.json'
        if not registry_file.exists():
            registry_file = project_root / 'element_maps' / domain / f'{page}.json'
        
        if not registry_file.exists():
            return jsonify({'error': f'Registry file not found: {domain}/{page}'}), 404
        
        # Use page name from URL with _page suffix for download filename (e.g., "home_page.json")
        download_filename = f'{page}_page.json'
        
        # Send file with proper headers for download
        response = send_file(
            str(registry_file),
            as_attachment=True,
            download_name=download_filename,
            mimetype='application/json'
        )
        
        # Explicitly set Content-Disposition header to ensure correct filename
        response.headers['Content-Disposition'] = f'attachment; filename="{download_filename}"'
        
        # Prevent browser caching
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"Error downloading registry JSON: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/registry', methods=['GET'])
def list_registries():
    """List all available registries (domains/pages)"""
    try:
        project_root = current_app.config['PROJECT_ROOT']
        registry_dir = project_root / 'element_maps'
        
        if not registry_dir.exists():
            return jsonify({'registries': []}), 200
        
        registries = []
        for domain_dir in registry_dir.iterdir():
            if domain_dir.is_dir() and not domain_dir.name.startswith('.'):
                domain = domain_dir.name
                for page_file in domain_dir.glob('*.json'):
                    page = page_file.stem.replace('_page', '')
                    registries.append({
                        'domain': domain,
                        'page': page,
                        'file': str(page_file)
                    })
        
        return jsonify({'registries': registries}), 200
        
    except Exception as e:
        print(f"Error listing registries: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/registry/<domain>/<page>', methods=['GET'])
def get_registry_for_page(domain, page):
    """Get registry content for a specific domain/page"""
    try:
        project_root = current_app.config['PROJECT_ROOT']
        registry_file = project_root / 'element_maps' / domain / f'{page}_page.json'
        
        if not registry_file.exists():
            return jsonify({'error': 'Registry not found'}), 404
        
        with open(registry_file, 'r') as f:
            registry = json.load(f)
        
        return jsonify(registry), 200
        
    except Exception as e:
        print(f"Error getting registry: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/registry/<domain>/<page>/element', methods=['PUT'])
def update_registry_element(domain, page):
    """Update an element in the registry"""
    try:
        data = request.get_json()
        element_key = data.get('element_key')
        new_selector = data.get('selector')
        
        if not element_key or not new_selector:
            return jsonify({'error': 'element_key and selector required'}), 400
        
        project_root = current_app.config['PROJECT_ROOT']
        registry_file = project_root / 'element_maps' / domain / f'{page}_page.json'
        
        if not registry_file.exists():
            return jsonify({'error': 'Registry not found'}), 404
        
        with open(registry_file, 'r') as f:
            registry = json.load(f)
        
        if element_key not in registry.get('elements', {}):
            return jsonify({'error': 'Element not found in registry'}), 404
        
        # Update the selector
        registry['elements'][element_key]['selector'] = new_selector
        registry['last_updated'] = datetime.now().isoformat() + 'Z'
        
        with open(registry_file, 'w') as f:
            json.dump(registry, f, indent=4)
        
        return jsonify({'success': True, 'message': 'Element updated'}), 200
        
    except Exception as e:
        print(f"Error updating registry element: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/registry/<domain>/<page>', methods=['DELETE'])
def delete_registry(domain, page):
    """Delete an entire registry (JSON file)"""
    try:
        project_root = current_app.config['PROJECT_ROOT']
        registry_file = project_root / 'element_maps' / domain / f'{page}_page.json'
        
        if not registry_file.exists():
            return jsonify({'error': 'Registry not found'}), 404
        
        # Create backup before deleting
        backup_file = registry_file.parent / f'{page}_page.json.deleted_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        import shutil
        shutil.copy(registry_file, backup_file)
        
        # Delete the registry file
        registry_file.unlink()
        
        return jsonify({
            'status': 'success',
            'message': f'Registry deleted successfully',
            'backup': str(backup_file)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/registry/<domain>/<page>/element', methods=['DELETE'])
def delete_registry_element(domain, page):
    """Delete an element from the registry"""
    try:
        data = request.get_json()
        element_key = data.get('element_key')
        
        if not element_key:
            return jsonify({'error': 'element_key required'}), 400
        
        project_root = current_app.config['PROJECT_ROOT']
        registry_file = project_root / 'element_maps' / domain / f'{page}_page.json'
        
        if not registry_file.exists():
            return jsonify({'error': 'Registry not found'}), 404
        
        with open(registry_file, 'r') as f:
            registry = json.load(f)
        
        if element_key not in registry.get('elements', {}):
            return jsonify({'error': 'Element not found in registry'}), 404
        
        # Delete the element
        del registry['elements'][element_key]
        registry['statistics']['total_elements'] = len(registry['elements'])
        registry['last_updated'] = datetime.now().isoformat() + 'Z'
        
        with open(registry_file, 'w') as f:
            json.dump(registry, f, indent=4)
        
        return jsonify({'success': True, 'message': 'Element deleted'}), 200
        
    except Exception as e:
        print(f"Error deleting registry element: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/parser/registry', methods=['GET'])
def get_registry_tree():
    """Load registry for tree viewer"""
    try:
        domain = request.args.get('domain')
        page = request.args.get('page')
        
        if not domain or not page:
            return jsonify({'error': 'Missing domain or page parameter'}), 400
        
        project_root = current_app.config['PROJECT_ROOT']
        registry = get_registry(str(project_root / 'element_maps'))
        
        element_map = registry.load_map(domain, page)
        
        if not element_map:
            return jsonify({'error': 'Registry not found'}), 404
        
        return jsonify(element_map), 200
        
    except Exception as e:
        import traceback
        print(f"Error loading registry: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/parser/registry', methods=['PUT'])
def update_registry_tree():
    """Save updated registry from tree viewer"""
    try:
        domain = request.args.get('domain')
        page = request.args.get('page')
        data = request.get_json()
        
        if not domain or not page:
            return jsonify({'error': 'Missing domain or page parameter'}), 400
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        project_root = current_app.config['PROJECT_ROOT']
        registry = get_registry(str(project_root / 'element_maps'))
        
        # Validate data structure
        if 'elements' not in data:
            return jsonify({'error': 'Invalid registry format'}), 400
        
        # Create backup before saving
        existing_map = registry.load_map(domain, page)
        if existing_map:
            backup_dir = project_root / 'element_maps' / domain / 'versions'
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = backup_dir / f'{page}_backup_{timestamp}.json'
            
            with open(backup_file, 'w') as f:
                json.dump(existing_map, f, indent=2)
            
            print(f"Created backup: {backup_file}")
        
        # Save updated registry
        data['updated_at'] = datetime.now().isoformat()
        data['updated_by'] = 'tree_editor'
        
        registry.save_map(domain, page, data)
        
        return jsonify({
            'success': True,
            'message': 'Registry updated successfully',
            'elements_count': len(data.get('elements', {}))
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error saving registry: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ============================================================================
# Excel API Routes (merged from REFACTOR/api/excel_routes.py)
# ============================================================================

@bp.route('/excel/upload', methods=['POST'])
def upload_excel():
    """
    Upload and validate Excel file, and optionally upload test files referenced in Excel.
    
    Expected form data:
    - file: Excel file (.xlsx or .xls)
    - test_files: (optional) Multiple test files for file upload steps
    
    Returns:
        JSON with excel_id, validation results, and test file upload status
    """
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file type. Expected .xlsx or .xls'}), 400
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Generate unique Excel ID
        excel_id = f"excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        # Create storage directories
        excel_files_dir = project_root / 'storage' / 'excel_files'
        excel_files_dir.mkdir(parents=True, exist_ok=True)
        
        metadata_dir = excel_files_dir / 'metadata'
        metadata_dir.mkdir(parents=True, exist_ok=True)
        
        # Save uploaded file
        excel_filename = f"{excel_id}.xlsx"
        excel_path = excel_files_dir / excel_filename
        file.save(str(excel_path))
        
        # Extract file upload paths from Excel
        file_upload_paths = extract_file_upload_paths_from_excel(excel_path)
        print(f"📁 Found {len(file_upload_paths)} file upload reference(s) in Excel: {file_upload_paths}")
        
        # Handle test file uploads if provided
        test_files_upload_result = {'uploaded': [], 'errors': [], 'referenced_paths': file_upload_paths, 'renamed': []}
        if 'test_files' in request.files:
            test_files = request.files.getlist('test_files')
            if test_files and any(f.filename for f in test_files):
                print(f"📤 Uploading {len([f for f in test_files if f.filename])} test file(s)...")
                print(f"📁 Excel references {len(file_upload_paths)} file path(s): {file_upload_paths}")
                upload_result = upload_test_files_to_server(test_files, project_root, referenced_paths=file_upload_paths)
                test_files_upload_result.update(upload_result)
        
        # Validate Excel file (with registry validation)
        validation_result = validate_excel_file(excel_path, project_root=project_root)
        
        # Save metadata
        metadata = {
            'excel_id': excel_id,
            'filename': file.filename,
            'saved_filename': excel_filename,
            'uploaded_at': datetime.now().isoformat(),
            'validation': validation_result,
            'file_path': str(excel_path.relative_to(project_root)),
            'file_upload_paths': file_upload_paths,
            'test_files_uploaded': test_files_upload_result
        }
        
        metadata_file = metadata_dir / f"{excel_id}.json"
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        # Return response
        response = {
            'success': validation_result['valid'],
            'excel_id': excel_id,
            'filename': file.filename,
            'validation': validation_result,
            'uploaded_at': metadata['uploaded_at'],
            'file_upload_paths': file_upload_paths,
            'test_files_upload': test_files_upload_result
        }
        
        if not validation_result['valid']:
            response['error'] = 'Excel file validation failed'
            response['validation_summary'] = get_validation_summary(validation_result)
            return jsonify(response), 400
        
        return jsonify(response), 200
        
    except Exception as e:
        import traceback
        print(f"Error uploading Excel file: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/generate', methods=['POST'])
def generate_from_excel():
    """
    Generate Playwright test from Excel file.
    
    Expected JSON:
    {
        "excel_id": "excel_...",
        "test_name": "optional_test_name"
    }
    
    Returns:
        JSON with generation status and test file info
    """
    try:
        data = request.get_json()
        excel_id = data.get('excel_id')
        
        if not excel_id:
            return jsonify({'error': 'excel_id required'}), 400
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load Excel metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get Excel file path
        excel_path = project_root / metadata['file_path']
        
        if not excel_path.exists():
            return jsonify({'error': f'Excel file not found: {excel_path}'}), 404
        
        # Generate test name
        test_name = data.get('test_name') or f"test_excel_{excel_id}"
        
        # Create output directory
        output_dir = project_root / 'storage' / 'excel_tests'
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_file = output_dir / f"{test_name}.spec.ts"
        
        # Track generation
        active_excel_generations[excel_id] = {
            'status': 'generating',
            'started_at': datetime.now().isoformat(),
            'excel_id': excel_id,
            'test_name': test_name
        }
        
        # Generate TypeScript Playwright code
        try:
            generation_result = generate_playwright_ts_from_excel(excel_path, output_file)
        except Exception as e:
            active_excel_generations[excel_id]['status'] = 'error'
            active_excel_generations[excel_id]['error'] = str(e)
            return jsonify({
                'success': False,
                'error': str(e),
                'excel_id': excel_id
            }), 500
        
        # Create execution ID for this Excel test
        execution_id = f"excel_exec_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{excel_id.split('_')[-1][:8]}"
        
        # Create execution metadata (similar to existing system)
        executions_dir = project_root / 'storage' / 'executions'
        executions_dir.mkdir(parents=True, exist_ok=True)
        
        execution_data = {
            'execution_id': execution_id,
            'excel_id': excel_id,
            'source': 'excel',
            'story': f"Excel test case: {metadata.get('filename', 'test_case.xlsx')}",
            'test_name': test_name,
            'test_file': str(output_file.relative_to(project_root)),
            'status': 'running',
            'created_at': datetime.now().isoformat(),
            'excel_metadata': {
                'excel_id': excel_id,
                'filename': metadata.get('filename'),
                'uploaded_at': metadata.get('uploaded_at'),
                'validation': metadata.get('validation', {})
            },
            'actions_taken': [],
            'screenshots': [],
            'playwright_validation': {
                'status': 'running',
                'test_running': True
            }
        }
        
        # Save initial execution data
        execution_file = executions_dir / f"{execution_id}.json"
        with open(execution_file, 'w') as f:
            json.dump(execution_data, f, indent=2)
        
        # Update metadata
        metadata['generated_test'] = {
            'test_name': test_name,
            'test_file': str(output_file.relative_to(project_root)),
            'generated_at': datetime.now().isoformat(),
            'generation_result': generation_result,
            'execution_id': execution_id
        }
        
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        # Update generation status
        active_excel_generations[excel_id]['status'] = 'completed'
        active_excel_generations[excel_id]['test_file'] = str(output_file.relative_to(project_root))
        active_excel_generations[excel_id]['execution_id'] = execution_id
        
        # Run test automatically in background thread
        def run_excel_test_background():
            try:
                # Import TypeScriptTestRunner
                try:
                    from validator.typescript_test_runner import TypeScriptTestRunner
                except ImportError as e:
                    print(f"⚠️ TypeScriptTestRunner not available - test will not be executed automatically: {e}")
                    import traceback
                    traceback.print_exc()
                    # Update execution with error
                    execution_data['status'] = 'error'
                    execution_data['error'] = f'TypeScriptTestRunner not available: {str(e)}'
                    execution_data['completed_at'] = datetime.now().isoformat()
                    with open(execution_file, 'w') as f:
                        json.dump(execution_data, f, indent=2)
                    return
                
                # TypeScriptTestRunner expects the full path to the test file
                runner = TypeScriptTestRunner(project_root)
                test_result = runner.run(str(output_file), execution_id=execution_id)
                
                # Update execution data with test results
                execution_data['status'] = 'completed' if test_result.get('status') == 'passed' else 'failed'
                execution_data['playwright_screenshots'] = test_result.get('screenshots', [])
                execution_data['screenshots'] = [s['filename'] for s in test_result.get('screenshots', [])]
                execution_data['playwright_validation'] = {
                    'status': test_result.get('status'),
                    'duration': test_result.get('duration'),
                    'assertions_passed': test_result.get('assertions_passed', 0),
                    'assertions_failed': test_result.get('assertions_failed', 0),
                    'test_file': test_result.get('test_file'),
                    'timestamp': test_result.get('timestamp'),
                    'stdout': test_result.get('stdout', ''),
                    'stderr': test_result.get('stderr', ''),
                    'exit_code': test_result.get('exit_code', 0),
                    'validation_mismatches': test_result.get('validation_mismatches', [])
                }
                execution_data['completed_at'] = datetime.now().isoformat()
                
                # Save updated execution data
                with open(execution_file, 'w') as f:
                    json.dump(execution_data, f, indent=2)
                
                # Update Excel metadata with execution results
                excel_metadata = metadata.copy()
                if 'test_executions' not in excel_metadata:
                    excel_metadata['test_executions'] = []
                
                excel_metadata['test_executions'].append({
                    'execution_id': execution_id,
                    'status': test_result.get('status'),
                    'duration': test_result.get('duration'),
                    'executed_at': datetime.now().isoformat()
                })
                excel_metadata['last_execution'] = excel_metadata['test_executions'][-1]
                
                with open(metadata_file, 'w') as f:
                    json.dump(excel_metadata, f, indent=2)
                
                print(f"✅ Excel test execution completed: {execution_id}")
                
            except Exception as e:
                import traceback
                print(f"❌ Error running Excel test: {e}")
                print(traceback.format_exc())
                
                # Update execution with error
                execution_data['status'] = 'error'
                execution_data['error'] = str(e)
                execution_data['completed_at'] = datetime.now().isoformat()
                
                with open(execution_file, 'w') as f:
                    json.dump(execution_data, f, indent=2)
        
        # Start background thread to run test
        thread = threading.Thread(target=run_excel_test_background)
        thread.daemon = True
        thread.start()
        
        # Return response with execution_id
        response = {
            'success': generation_result.get('success', True),
            'excel_id': excel_id,
            'execution_id': execution_id,
            'test_name': test_name,
            'test_file': str(output_file.relative_to(project_root)),
            'rows_processed': generation_result.get('rows_processed', 0),
            'generated_at': datetime.now().isoformat(),
            'test_running': True,
            'results_url': f'/results/{execution_id}'
        }
        
        if generation_result.get('errors'):
            response['warnings'] = generation_result['errors']
            # If generation failed, include error message
            if not generation_result.get('success', True):
                response['error'] = '; '.join(generation_result['errors']) if generation_result['errors'] else 'Generation failed'
        
        return jsonify(response), 200
        
    except Exception as e:
        import traceback
        print(f"Error generating test from Excel: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/status', methods=['GET'])
def get_excel_status(excel_id):
    """
    Get Excel file generation status.
    
    Returns:
        JSON with Excel file status and metadata
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get generation status if available
        generation_status = active_excel_generations.get(excel_id, {})
        
        response = {
            'excel_id': excel_id,
            'filename': metadata.get('filename'),
            'uploaded_at': metadata.get('uploaded_at'),
            'validation': metadata.get('validation', {}),
            'generation_status': generation_status.get('status', 'not_started')
        }
        
        if 'generated_test' in metadata:
            response['generated_test'] = metadata['generated_test']
        
        if 'error' in generation_status:
            response['error'] = generation_status['error']
        
        return jsonify(response), 200
        
    except Exception as e:
        import traceback
        print(f"Error getting Excel status: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/template', methods=['GET'])
def download_excel_template():
    """
    Download Excel template file.
    
    Query params:
    - include_examples: true/false (default: true)
    
    Returns:
        Excel template file download
    """
    try:
        include_examples = request.args.get('include_examples', 'true').lower() == 'true'
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Generate template using project root
        template_path = get_template_path("test_case_template.xlsx", project_root=project_root)
        generate_excel_template(template_path, include_examples=include_examples)
        
        if not template_path.exists():
            return jsonify({'error': 'Failed to generate template'}), 500
        
        return send_file(
            str(template_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='test_case_template.xlsx'
        )
        
    except Exception as e:
        import traceback
        print(f"Error downloading template: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/metadata', methods=['GET'])
def get_excel_metadata(excel_id):
    """
    Get Excel file metadata.
    
    Returns:
        JSON with Excel file metadata
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        return jsonify(metadata), 200
        
    except Exception as e:
        import traceback
        print(f"Error getting Excel metadata: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/download', methods=['GET'])
def download_excel_file(excel_id):
    """
    Download uploaded Excel file.
    
    Returns:
        Excel file download
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        excel_path = project_root / metadata['file_path']
        
        if not excel_path.exists():
            return jsonify({'error': f'Excel file not found: {excel_path}'}), 404
        
        return send_file(
            str(excel_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=metadata.get('filename', 'test_case.xlsx')
        )
        
    except Exception as e:
        import traceback
        print(f"Error downloading Excel file: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/generate-ts', methods=['POST'])
def generate_ts_from_excel():
    """
    Generate TypeScript Playwright test from Excel file.
    
    Expected JSON:
    {
        "excel_id": "excel_...",
        "test_name": "optional_test_name"
    }
    
    Returns:
        JSON with generation status and test file info
    """
    try:
        data = request.get_json()
        excel_id = data.get('excel_id')
        
        if not excel_id:
            return jsonify({'error': 'excel_id required'}), 400
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load Excel metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get Excel file path
        excel_path = project_root / metadata['file_path']
        
        if not excel_path.exists():
            return jsonify({'error': f'Excel file not found: {excel_path}'}), 404
        
        # Generate test name
        test_name = data.get('test_name') or f"test_excel_{excel_id}"
        
        # Create output directory
        output_dir = project_root / 'storage' / 'excel_tests'
        output_dir.mkdir(parents=True, exist_ok=True)
        
        output_file = output_dir / f"{test_name}.spec.ts"  # .spec.ts extension
        
        # Generate TypeScript code
        try:
            generation_result = generate_playwright_ts_from_excel(excel_path, output_file)
            
            # Check if file was actually created (validation warnings are OK)
            if not output_file.exists():
                error_msg = generation_result.get('error', 'Unknown error')
                if generation_result.get('errors'):
                    error_msg += ': ' + '; '.join(generation_result['errors'][:3])
                return jsonify({
                    'success': False,
                    'error': error_msg,
                    'excel_id': excel_id
                }), 500
        except Exception as e:
            return jsonify({
                'success': False,
                'error': str(e),
                'excel_id': excel_id
            }), 500
        
        # Update metadata with TypeScript test info
        if 'generated_test_ts' not in metadata:
            metadata['generated_test_ts'] = {}
        
        metadata['generated_test_ts'] = {
            'test_name': test_name,
            'test_file': str(output_file.relative_to(project_root)),
            'generated_at': datetime.now().isoformat(),
            'generation_result': generation_result
        }
        
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        # Check if we should run the test (default: True)
        run_test = data.get('run_test', True)
        
        response_data = {
            'success': True,
            'test_file': str(output_file.relative_to(project_root)),
            'test_name': test_name,
            'excel_id': excel_id,
            'download_url': f'/api/excel/{excel_id}/test-ts',
            'zip_download_url': f'/api/excel/{excel_id}/test-ts-zip',
            'generation_result': generation_result
        }
        
        # Run test in background if requested
        if run_test:
            # Find or create execution_id for this Excel file
            # Check if there's an existing execution for this Excel
            executions_dir = project_root / 'storage' / 'executions'
            execution_id = None
            
            # Look for existing execution with this excel_id
            if executions_dir.exists():
                for exec_file in executions_dir.glob('*.json'):
                    try:
                        with open(exec_file, 'r') as f:
                            exec_data = json.load(f)
                        if exec_data.get('excel_id') == excel_id:
                            execution_id = exec_data.get('execution_id') or exec_file.stem
                            break
                    except:
                        continue
            
            # Create new execution if not found
            if not execution_id:
                execution_id = f"excel_exec_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
                execution_file = executions_dir / f'{execution_id}.json'
                executions_dir.mkdir(parents=True, exist_ok=True)
                
                # Create execution data
                exec_data = {
                    'execution_id': execution_id,
                    'excel_id': excel_id,
                    'source': 'excel',
                    'status': 'running',
                    'created_at': datetime.now().isoformat(),
                    'test_file': str(output_file.relative_to(project_root)),
                    'test_name': test_name
                }
                
                with open(execution_file, 'w') as f:
                    json.dump(exec_data, f, indent=2)
            
            # Run test in background thread
            def run_test_background():
                try:
                    from validator.typescript_test_runner import TypeScriptTestRunner
                    
                    runner = TypeScriptTestRunner(project_root)
                    test_result = runner.run(str(output_file), execution_id)
                    
                    # Update execution with test results
                    execution_file = executions_dir / f'{execution_id}.json'
                    if execution_file.exists():
                        with open(execution_file, 'r') as f:
                            exec_data = json.load(f)
                        
                        exec_data['playwright_validation'] = {
                            'status': test_result.get('status'),
                            'duration': test_result.get('duration'),
                            'assertions_passed': test_result.get('assertions_passed', 0),
                            'assertions_failed': test_result.get('assertions_failed', 0),
                            'test_file': test_result.get('test_file'),
                            'timestamp': test_result.get('timestamp'),
                            'stdout': test_result.get('stdout', ''),
                            'stderr': test_result.get('stderr', ''),
                            'exit_code': test_result.get('exit_code', 0)
                        }
                        exec_data['playwright_screenshots'] = test_result.get('screenshots', [])
                        # Set status based on test result: 'failed' if test failed, 'completed' if passed
                        test_status = test_result.get('status', 'unknown')
                        if test_status == 'failed' or test_result.get('exit_code', 0) != 0:
                            exec_data['status'] = 'failed'
                        else:
                            exec_data['status'] = 'completed'
                        exec_data['completed_at'] = datetime.now().isoformat()
                        
                        with open(execution_file, 'w') as f:
                            json.dump(exec_data, f, indent=2)
                        
                        print(f"✅ TypeScript test execution completed for {execution_id}")
                except Exception as e:
                    print(f"Error running TypeScript test in background: {e}")
                    import traceback
                    traceback.print_exc()
                    
                    # Even on error, try to collect any screenshots that were created
                    execution_file = executions_dir / f'{execution_id}.json'
                    if execution_file.exists():
                        try:
                            from validator.typescript_test_runner import TypeScriptTestRunner
                            runner = TypeScriptTestRunner(project_root)
                            # Collect screenshots from disk (test may have generated some before error)
                            screenshots = runner._collect_screenshots(time.time(), 0)
                            
                            with open(execution_file, 'r') as f:
                                exec_data = json.load(f)
                            
                            exec_data['playwright_screenshots'] = screenshots
                            exec_data['status'] = 'error'
                            exec_data['error'] = str(e)
                            exec_data['completed_at'] = datetime.now().isoformat()
                            
                            with open(execution_file, 'w') as f:
                                json.dump(exec_data, f, indent=2)
                            
                            print(f"✅ Collected {len(screenshots)} screenshots despite error")
                        except Exception as collect_error:
                            print(f"⚠️  Could not collect screenshots: {collect_error}")
            
            thread = threading.Thread(target=run_test_background)
            thread.daemon = True
            thread.start()
            
            response_data['execution_id'] = execution_id
            response_data['results_url'] = f'/results/{execution_id}'
            response_data['test_running'] = True
            response_data['message'] = 'TypeScript test generation completed. Test is running in background. Screenshots will be available shortly.'
        
        return jsonify(response_data)
        
    except Exception as e:
        import traceback
        print(f"Error generating TypeScript test: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/test-ts', methods=['GET'])
def download_ts_test(excel_id):
    """
    Download generated TypeScript Playwright test file.
    
    Returns:
        TypeScript test file download
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        if 'generated_test_ts' not in metadata:
            return jsonify({'error': 'TypeScript test not generated yet'}), 404
        
        test_file_path = project_root / metadata['generated_test_ts']['test_file']
        
        if not test_file_path.exists():
            return jsonify({'error': f'TypeScript test file not found: {test_file_path}'}), 404
        
        return send_file(
            str(test_file_path),
            mimetype='text/typescript',
            as_attachment=True,
            download_name=metadata['generated_test_ts']['test_name'] + '.spec.ts'
        )
        
    except Exception as e:
        import traceback
        print(f"Error downloading TypeScript test file: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/test-ts-zip', methods=['GET'])
def download_ts_test_zip(excel_id):
    """
    Download TypeScript test file (.spec.ts), package.json, README, and all required registry JSON files bundled as a zip file (excludes .env for security)
    
    Returns:
        Zip file download
    """
    try:
        from flask import send_file
        from io import BytesIO
        import zipfile
        import re
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        if 'generated_test_ts' not in metadata:
            return jsonify({'error': 'TypeScript test not generated yet'}), 404
        
        test_file_path = project_root / metadata['generated_test_ts']['test_file']
        test_name = metadata['generated_test_ts']['test_name']
        
        if not test_file_path.exists():
            return jsonify({'error': f'TypeScript test file not found: {test_file_path}'}), 404
        
        # Read TypeScript test file to extract REGISTRY_PATHS
        with open(test_file_path, 'r') as f:
            ts_content = f.read()
        
        # Extract REGISTRY_PATHS from TypeScript file
        registry_paths = []
        # Match: const REGISTRY_PATHS = [\n    'path1',\n    'path2',\n]
        match = re.search(r"const\s+REGISTRY_PATHS\s*=\s*\[(.*?)\]", ts_content, re.DOTALL)
        if match:
            paths_str = match.group(1)
            # Extract all quoted strings
            path_matches = re.findall(r"['\"]([^'\"]+)['\"]", paths_str)
            registry_paths = [p.strip() for p in path_matches if p.strip()]
        
        # Extract file upload paths from TypeScript file
        # Look for patterns like: path.resolve(projectRoot, 'storage/test_files/filename.tsv')
        test_file_paths = []
        file_path_pattern = r"path\.resolve\(projectRoot,\s*['\"]([^'\"]+)['\"]\)"
        file_path_matches = re.findall(file_path_pattern, ts_content)
        for file_path in file_path_matches:
            file_path = file_path.strip()
            if file_path and file_path not in test_file_paths:
                test_file_paths.append(file_path)
        
        # Get Excel filename for README and ZIP
        excel_filename = metadata.get('filename', 'test_case.xlsx')
        
        # Create package.json content (with xlsx for Excel reading)
        package_json_content = '''{
  "name": "playwright-test",
  "version": "1.0.0",
  "scripts": {
    "test": "playwright test"
  },
  "dependencies": {
    "@playwright/test": "^1.40.0",
    "dotenv": "^16.0.0",
    "xlsx": "^0.18.5"
  }
}'''
        
        # Create README content
        test_files_note = ""
        if test_file_paths:
            test_files_note = f"\n- `storage/test_files/` - Test data files for file upload steps ({len(test_file_paths)} file(s))"
        
        readme_content = f'''# {test_name} - Playwright TypeScript Test

## Setup Instructions

1. Extract this zip file
2. The `.env` file is already included with credentials
3. Run: `npm install`
4. Run: `npx playwright install chromium`
5. Run: `npx playwright test {test_name}.spec.ts --headed`

## Files Included
- `{test_name}.spec.ts` - Main test file
- `package.json` - Dependencies
- `generate_totp.py` - Python script for TOTP generation (called from TypeScript)
- `element_maps/` - JSON registry files with element XPath mappings{test_files_note}

## Important
- **`.env` file is included** with the package (contains TOTP secrets and credentials)
- Registry JSON files are included in the `element_maps/` directory structure
- Test data files are included in the `storage/test_files/` directory structure (if file uploads are used)
- The test uses `pyotp` via Python script for TOTP generation to ensure consistency with Python tests

## Notes
- Screenshots saved to `storage/screenshots/`
- Test uses registry-based element lookup (no hard-coded XPaths)
- File paths are resolved relative to the test file location (goes up 2 levels to project root)
'''
        
        # Create zip file in memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add TypeScript test file
            ts_filename = f'{test_name}.spec.ts'
            zip_file.write(test_file_path, ts_filename)
            print(f"✅ Added TypeScript test file to zip: {ts_filename}")
            
            # Note: .env file is NOT included for security reasons
            
            # Add package.json
            zip_file.writestr('package.json', package_json_content)
            print(f"✅ Added package.json to zip")
            
            # Add README.md
            zip_file.writestr('README.md', readme_content)
            print(f"✅ Added README.md to zip")
            
            # Add generate_totp.py script (required for TOTP generation in TypeScript tests)
            totp_script_path = project_root / 'Test' / 'generate_totp.py'
            if totp_script_path.exists():
                zip_file.write(totp_script_path, 'generate_totp.py')
                print(f"✅ Added generate_totp.py to zip")
            else:
                print(f"⚠️  generate_totp.py not found at {totp_script_path}")
            
            # Add .env file from Test directory
            env_file_path = project_root / 'Test' / '.env'
            if env_file_path.exists():
                zip_file.write(env_file_path, '.env')
                print(f"✅ Added .env file")
            else:
                print(f"⚠️  .env file not found at {env_file_path}")
            
            # Add all registry JSON files with proper directory structure
            for registry_path in registry_paths:
                # registry_path is like 'element_maps/domain/page_page.json'
                full_registry_path = project_root / registry_path
                
                if full_registry_path.exists():
                    # Preserve the full path structure as the test expects it
                    zip_path = registry_path
                    zip_file.write(full_registry_path, zip_path)
                    print(f"✅ Added registry file to zip: {zip_path}")
                else:
                    print(f"⚠️  Registry file not found: {full_registry_path}")
            
            # Add test files for file upload steps
            if test_file_paths:
                for file_path in test_file_paths:
                    # file_path is like 'storage/test_files/dcpagain-icdc_file.tsv'
                    full_file_path = project_root / file_path
                    
                    if full_file_path.exists():
                        # Preserve the full path structure as the test expects it
                        zip_path = file_path
                        zip_file.write(full_file_path, zip_path)
                        print(f"✅ Added test file to zip: {zip_path}")
                    else:
                        print(f"⚠️  Test file not found: {full_file_path}")
            else:
                print("ℹ️  No file upload paths found in test file")
            
            # Add Excel file (source file for credentials and expected results)
            excel_path = project_root / metadata['file_path']
            if excel_path.exists():
                zip_file.write(excel_path, excel_filename)
                print(f"✅ Added Excel file to zip: {excel_filename}")
            else:
                print(f"⚠️  Excel file not found: {excel_path}")
        
        # Verify zip contents
        zip_buffer.seek(0)
        with zipfile.ZipFile(zip_buffer, 'r') as verify_zip:
            file_list = verify_zip.namelist()
            print(f"📦 Zip contains {len(file_list)} files:")
            for name in file_list:
                print(f"   - {name}")
        
        zip_buffer.seek(0)
        
        # Create zip filename
        zip_filename = f'{test_name}_typescript_complete.zip'
        
        # Send zip file
        response = send_file(
            zip_buffer,
            as_attachment=True,
            download_name=zip_filename,
            mimetype='application/zip'
        )
        
        response.headers['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        import traceback
        print(f"Error creating TypeScript zip file: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/test', methods=['GET'])
def download_excel_generated_test(excel_id):
    """
    Download generated Playwright test file.
    
    Returns:
        TypeScript test file download
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        if 'generated_test' not in metadata:
            return jsonify({'error': 'Test not generated yet'}), 404
        
        test_file_path = project_root / metadata['generated_test']['test_file']
        
        if not test_file_path.exists():
            return jsonify({'error': f'Test file not found: {test_file_path}'}), 404
        
        return send_file(
            str(test_file_path),
            mimetype='text/typescript',
            as_attachment=True,
            download_name=metadata['generated_test']['test_name'] + '.spec.ts'
        )
        
    except Exception as e:
        import traceback
        print(f"Error downloading test file: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/registry/compare', methods=['GET'])
def compare_excel_with_registry(excel_id):
    """
    Compare Excel file elements with registry to find new/updated elements.
    
    Returns:
        JSON with new_elements, updated_elements, unchanged_elements
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get Excel file path
        excel_path = project_root / metadata['file_path']
        
        if not excel_path.exists():
            return jsonify({'error': f'Excel file not found: {excel_path}'}), 404
        
        # Extract elements from Excel
        elements = extract_elements_from_excel(excel_path)
        
        # Load registry
        from utils.element_registry import get_registry
        registry = get_registry()
        
        # Compare with registry
        comparison = compare_with_registry(elements, registry)
        
        return jsonify({
            'success': True,
            'comparison': comparison,
            'excel_id': excel_id
        }), 200
    
    except Exception as e:
        import traceback
        print(f"Error comparing Excel with registry: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/registry/update', methods=['POST'])
def update_registry_from_excel(excel_id):
    """
    Update registry with elements from Excel file.
    
    Expected JSON:
    {
        "new_elements": [...],  # Elements to add
        "updated_elements": [...]  # Elements to update
    }
    
    Returns:
        JSON with update status
    """
    try:
        data = request.get_json()
        new_elements = data.get('new_elements', [])
        updated_elements = data.get('updated_elements', [])
        
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        # Load registry
        from utils.element_registry import get_registry
        registry = get_registry()
        
        added_count = 0
        updated_count = 0
        errors = []
        
        # Add new elements
        for elem in new_elements:
            try:
                domain = elem['domain']
                page = elem['page']
                element_name = elem['element_name']
                xpath = elem['xpath']
                
                # Load or create element map
                element_map = registry.load_map(domain, page)
                if not element_map:
                    from datetime import datetime
                    element_map = {
                        "page": page,
                        "url": elem['url'],
                        "version": "1.0",
                        "timestamp": datetime.now().isoformat() + "Z",
                        "elements": {},
                        "id_index": {},
                        "statistics": {
                            "total_elements": 0,
                            "parsed_elements": 0,
                            "discovered_elements": 0
                        }
                    }
                
                # Add element
                if element_name not in element_map['elements']:
                    element_map['elements'][element_name] = {
                        'xpath': xpath,
                        'selector': xpath,
                        'element_id': registry._generate_element_id(element_name, xpath),
                        'usage_count': 0,
                        'last_used': None,
                        'source': 'excel',
                        'object_type': elem.get('object_type', ''),
                        'action': elem.get('action', '')
                    }
                    element_map['statistics']['total_elements'] = len(element_map['elements'])
                    registry.save_map(domain, page, element_map)
                    added_count += 1
            except Exception as e:
                errors.append(f"Failed to add {elem.get('element_name', 'unknown')}: {str(e)}")
        
        # Update existing elements
        for elem in updated_elements:
            try:
                domain = elem['domain']
                page = elem['page']
                element_name = elem['element_name']
                new_xpath = elem['new_xpath']
                
                # Load element map
                element_map = registry.load_map(domain, page)
                if not element_map:
                    errors.append(f"Registry not found for {domain}/{page}")
                    continue
                
                # Update element
                if element_name in element_map['elements']:
                    element_map['elements'][element_name]['xpath'] = new_xpath
                    element_map['elements'][element_name]['selector'] = new_xpath
                    element_map['elements'][element_name]['source'] = 'excel_updated'
                    registry.save_map(domain, page, element_map)
                    updated_count += 1
                else:
                    errors.append(f"Element {element_name} not found in registry")
            except Exception as e:
                errors.append(f"Failed to update {elem.get('element_name', 'unknown')}: {str(e)}")
        
        return jsonify({
            'success': True,
            'added_count': added_count,
            'updated_count': updated_count,
            'errors': errors,
            'excel_id': excel_id
        }), 200
    
    except Exception as e:
        import traceback
        print(f"Error updating registry from Excel: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@bp.route('/excel/<excel_id>/steps', methods=['GET'])
def get_excel_steps(excel_id):
    """
    Get test steps from Excel file.
    
    Returns:
        JSON with steps array
    """
    try:
        project_root = current_app.config.get('PROJECT_ROOT', Path.cwd())
        
        # Load metadata
        metadata_dir = project_root / 'storage' / 'excel_files' / 'metadata'
        metadata_file = metadata_dir / f"{excel_id}.json"
        
        if not metadata_file.exists():
            return jsonify({'error': f'Excel file not found: {excel_id}'}), 404
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        # Get Excel file path
        excel_path = project_root / metadata['file_path']
        
        if not excel_path.exists():
            return jsonify({'error': f'Excel file not found: {excel_path}'}), 404
        
        # Read Excel file and extract steps
        import pandas as pd
        df = pd.read_excel(excel_path)
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        steps = []
        for idx, row in df.iterrows():
            step = str(row.get('step', idx + 1)).strip()
            url = str(row.get('url', '')).strip() if pd.notna(row.get('url')) else ''
            xpath = str(row.get('xpath', '')).strip() if pd.notna(row.get('xpath')) else ''
            action = str(row.get('action', '')).strip().lower() if pd.notna(row.get('action')) else ''
            object_type_raw = row.get('object_type', '')
            object_type = str(object_type_raw).strip() if pd.notna(object_type_raw) and object_type_raw != '' else ''
            text_value = str(row.get('text_value', '')).strip() if pd.notna(row.get('text_value')) else ''
            # Handle wait_time - convert NaN to None for JSON compatibility
            wait_time_raw = row.get('wait_time', None)
            if pd.isna(wait_time_raw):
                wait_time = None
            else:
                try:
                    wait_time = float(wait_time_raw) if wait_time_raw is not None else None
                except (ValueError, TypeError):
                    wait_time = None
            
            functions = str(row.get('functions', '')).strip() if pd.notna(row.get('functions')) else ''
            is_optional = str(row.get('optional', '')).strip().lower() in ['true', 'yes', '1', 'y']
            
            steps.append({
                'step': step,
                'url': url if url and url.upper() != 'N/A' else None,
                'xpath': xpath if xpath and xpath.upper() != 'N/A' else None,
                'action': action,
                'object_type': object_type,
                'text_value': text_value if text_value else None,
                'wait_time': wait_time,
                'functions': functions if functions else None,
                'optional': is_optional
            })
        
        return jsonify({
            'success': True,
            'steps': steps,
            'excel_id': excel_id
        }), 200
    
    except Exception as e:
        import traceback
        print(f"Error getting Excel steps: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

